"""
app.py — Main Flask application: routes, views, and app factory
"""
import json
import logging
import os
import sys
import threading
import uuid
from datetime import datetime, timedelta
from io import BytesIO

from flask import (
    Flask, Response, abort, flash, jsonify, redirect,
    render_template, request, send_file, session, url_for,
)

from config import Config
from extensions import db
from models import Application, Job, Notification, ScanLog, ScraperSource, Setting

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Thread-safe state for the "Check Now" email progress window
# ---------------------------------------------------------------------------
_email_check_lock  = threading.Lock()
_email_check_state = {"running": False, "log": [], "finished": True}


def _resource_dir() -> str:
    """
    Directory that contains the bundled read-only assets (templates/, static/).
    • PyInstaller frozen  → sys._MEIPASS  (the extraction temp dir)
    • Development         → directory of this file
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


def _data_dir() -> str:
    """
    User-writable directory for the database, logs, and .env.
    • Frozen → directory of the .exe (install folder chosen by the user)
    • Development → directory of this file
    """
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

# Capture all log output in-memory so it can be viewed in the browser
from server_log import init_server_log
init_server_log()


# ---------------------------------------------------------------------------
# App factory
# ---------------------------------------------------------------------------

def create_app(config_class=Config):
    _res = _resource_dir()
    _dat = _data_dir()

    app = Flask(
        __name__,
        template_folder=os.path.join(_res, "templates"),
        static_folder=os.path.join(_res, "static"),
    )
    app.config.from_object(config_class)

    # When frozen, override the DB URI to use an absolute path in the install
    # folder so the database is always next to the executable, regardless of
    # the working directory the OS uses when launching the shortcut.
    if getattr(sys, "frozen", False):
        _db = os.path.join(_dat, "job_tracker.db").replace("\\", "/")
        app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{_db}"

    db.init_app(app)

    with app.app_context():
        db.create_all()
        # Safe schema migrations (no-op after first run)
        for _sql in [
            "ALTER TABLE notifications ADD COLUMN is_archived BOOLEAN NOT NULL DEFAULT 0",
            "ALTER TABLE jobs ADD COLUMN job_duration VARCHAR(50)",
            "ALTER TABLE jobs ADD COLUMN company_address VARCHAR(500)",
            "ALTER TABLE jobs ADD COLUMN company_phone VARCHAR(100)",
            "ALTER TABLE scraper_sources ADD COLUMN health_status VARCHAR(20)",
            "ALTER TABLE scraper_sources ADD COLUMN health_checked_at DATETIME",
            "ALTER TABLE notifications ADD COLUMN is_acknowledged BOOLEAN DEFAULT 0",
        ]:
            try:
                db.session.execute(db.text(_sql))
                db.session.commit()
            except Exception:
                db.session.rollback()
        _seed_default_settings(app)
        # Seed built-in scraper rows (safe to call multiple times)
        from scrapers import registry as scraper_registry
        scraper_registry.seed_db(app)

    # Register Jinja2 globals
    @app.context_processor
    def inject_globals():
        unread = Notification.query.filter_by(is_read=False, is_archived=False).count()
        recent_notifs = (
            Notification.query
            .filter_by(is_archived=False)
            .order_by(Notification.created_at.desc())
            .limit(8)
            .all()
        )
        last_scan = ScanLog.query.order_by(ScanLog.scan_date.desc()).first()
        unemp_enabled = Setting.get("unemployment_enabled", "false") == "true"
        unemp_url     = Setting.get("unemployment_url", "") if unemp_enabled else ""
        unemp_state   = Setting.get("unemployment_state", "") if unemp_enabled else ""
        try:
            _pdc = {
                'range1_days':  int(Setting.get('posted_date_range1_days',  '7')  or '7'),
                'range1_color': Setting.get('posted_date_range1_color', '#28a745') or '#28a745',
                'range2_days':  int(Setting.get('posted_date_range2_days',  '30') or '30'),
                'range2_color': Setting.get('posted_date_range2_color', '#fd7e14') or '#fd7e14',
                'range3_color': '#adb5bd',
            }
        except Exception:
            _pdc = {'range1_days': 7, 'range1_color': '#28a745',
                    'range2_days': 30, 'range2_color': '#fd7e14', 'range3_color': '#adb5bd'}
        return dict(unread_count=unread, recent_notifications=recent_notifs,
                    last_scan=last_scan,
                    unemp_enabled=unemp_enabled,
                    unemp_url=unemp_url,
                    unemp_state=unemp_state,
                    posted_date_colors=_pdc)

    _register_routes(app)
    return app


def _seed_default_settings(app):
    """Insert default settings rows on first run."""
    defaults = [
        ("applicant_name", Config.APPLICANT_NAME, "Your full name"),
        ("applicant_email", Config.APPLICANT_EMAIL, "Contact email"),
        ("applicant_phone", Config.APPLICANT_PHONE, "Phone number"),
        ("applicant_location", Config.APPLICANT_LOCATION, "City / Country"),
        ("years_experience", Config.YEARS_EXPERIENCE, "Years of experience"),
        ("applicant_linkedin", Config.APPLICANT_LINKEDIN, "LinkedIn profile URL"),
        ("applicant_github", Config.APPLICANT_GITHUB, "GitHub profile URL"),
        ("resume_path", Config.RESUME_PATH, "Path to your resume file"),
        ("min_match_score", str(Config.MIN_MATCH_SCORE), "Minimum score to keep a job (0-100)"),
        ("required_skills", ",".join(Config.REQUIRED_SKILLS), "Comma-separated required skills"),
        ("preferred_skills", ",".join(Config.PREFERRED_SKILLS), "Comma-separated preferred skills"),
        ("follow_up_days", str(Config.FOLLOW_UP_DAYS), "Days before follow-up reminder"),
        ("scan_morning", Config.SCAN_TIME_MORNING, "Morning scan time (HH:MM)"),
        ("scan_evening", Config.SCAN_TIME_EVENING, "Evening scan time (HH:MM)"),
        ("timezone", Config.TIMEZONE, "Timezone for scheduled scans"),
        ("use_ai_cover_letter", "false", "Use OpenAI for cover letter generation"),
        ("applicant_location_miles", "50", "Miles radius for in-person job filtering"),
        ("notif_auto_archive_enabled", "false", "Auto-archive old notifications"),
        ("notif_auto_archive_days", "30", "Days before a notification is auto-archived"),
        ("notif_auto_delete_enabled", "false", "Auto-delete old notifications"),
        ("notif_auto_delete_days", "60", "Days before a notification is auto-deleted"),
        ("scan_auto_enabled", "true", "Enable automated scheduled scans"),
        ("scan_frequency", "daily", "Scan frequency: daily, weekly, or monthly"),
        ("scan_times", "08:00,20:00", "Comma-separated scan times (HH:MM)"),
        ("scan_weekdays", "0,1,2,3,4", "Days of week for weekly scans (0=Mon … 6=Sun)"),
        ("scan_monthdays", "1", "Day(s) of month for monthly scans"),
        ("job_duration_filter", "", "Comma-separated accepted job durations (empty = all)"),
        ("discovered_skills", "{}", "JSON map of skills discovered during scans"),
        ("use_resume", "false", "Use resume file to source skills instead of manual fields"),
        ("resume_required_skills", "", "Resume skills weighted as required (+20 pts each)"),
        ("resume_preferred_skills", "", "Resume skills weighted as preferred (+8 pts each)"),
        ("show_tray_icon", "true", "Show system tray icon when the background daemon is running"),
        ("app_port", "5000", "Port the web app runs on (used by tray icon Open link)"),
        ("daemon_name", "JobTrackerDaemon", "Process name shown in Windows Task Manager for the background daemon"),
        ("cal_feed_token", uuid.uuid4().hex, "Secret token for the ICS calendar feed URL"),
        ("posted_date_range1_days",  "7",        "Days threshold for first posted-date color range"),
        ("posted_date_range1_color", "#28a745",  "Color for recently posted jobs (range 1)"),
        ("posted_date_range2_days",  "30",       "Days threshold for second posted-date color range"),
        ("posted_date_range2_color", "#fd7e14",  "Color for moderately old job postings (range 2)"),
        ("scan_timeout_minutes",     "30",       "Max scan duration in minutes; daemon restarts on timeout (0 = disabled)"),
    ]
    for key, value, description in defaults:
        if not Setting.query.filter_by(key=key).first():
            db.session.add(Setting(key=key, value=value, description=description))
    db.session.commit()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _settings_as_config():
    """Build a simple namespace from DB settings (used in services)."""
    class Cfg:
        pass
    cfg = Cfg()
    rows = Setting.query.all()
    mapping = {r.key: r.value for r in rows}

    cfg.APPLICANT_NAME = mapping.get("applicant_name", "")
    cfg.APPLICANT_EMAIL = mapping.get("applicant_email", "")
    cfg.APPLICANT_PHONE = mapping.get("applicant_phone", "")
    cfg.APPLICANT_LOCATION = mapping.get("applicant_location", "")
    cfg.YEARS_EXPERIENCE = mapping.get("years_experience", "5")
    cfg.APPLICANT_LINKEDIN = mapping.get("applicant_linkedin", "")
    cfg.APPLICANT_GITHUB = mapping.get("applicant_github", "")
    cfg.RESUME_PATH = mapping.get("resume_path", "")

    cfg.SMTP_HOST = Config.SMTP_HOST
    cfg.SMTP_PORT = Config.SMTP_PORT
    cfg.SMTP_USER = Config.SMTP_USER
    cfg.SMTP_PASSWORD = Config.SMTP_PASSWORD
    cfg.FROM_EMAIL = Config.FROM_EMAIL
    cfg.NOTIFY_EMAIL = Config.NOTIFY_EMAIL

    cfg.OPENAI_API_KEY = Config.OPENAI_API_KEY
    cfg.OPENAI_MODEL = Config.OPENAI_MODEL

    return cfg


# ---------------------------------------------------------------------------
# Notification cleanup helper
# ---------------------------------------------------------------------------

def _apply_notification_cleanup():
    """Archive or delete notifications that have exceeded configured age limits."""
    now = datetime.utcnow()

    if Setting.get("notif_auto_archive_enabled", "false") == "true":
        try:
            days = int(Setting.get("notif_auto_archive_days", "30"))
            cutoff = now - timedelta(days=days)
            Notification.query.filter(
                Notification.created_at < cutoff,
                Notification.is_archived == False,  # noqa: E712
            ).update({"is_archived": True})
            db.session.commit()
        except Exception:
            db.session.rollback()

    if Setting.get("notif_auto_delete_enabled", "false") == "true":
        try:
            days = int(Setting.get("notif_auto_delete_days", "60"))
            cutoff = now - timedelta(days=days)
            Notification.query.filter(
                Notification.created_at < cutoff
            ).delete()
            db.session.commit()
        except Exception:
            db.session.rollback()


# ---------------------------------------------------------------------------
# Route registration
# ---------------------------------------------------------------------------

def _register_routes(app):

    # ------------------------------------------------------------------
    # Dashboard
    # ------------------------------------------------------------------

    @app.route("/")
    def dashboard():
        from analytics import key_metrics, upcoming_followups, pipeline_counts
        metrics = key_metrics()
        followups = upcoming_followups(5)
        pipeline = pipeline_counts()
        recent_jobs = (
            Job.query
            .filter(Job.status != "archived")
            .order_by(Job.match_score.desc(), Job.found_date.desc())
            .all()
        )
        scan_logs = ScanLog.query.order_by(ScanLog.scan_date.desc()).limit(20).all()

        # Score tier thresholds derived from MIN_MATCH_SCORE setting
        min_score = int(Setting.get("min_match_score", 50))
        tier = (100 - min_score) / 3
        score_t1 = min_score + tier        # grey → light-blue boundary
        score_t2 = min_score + 2 * tier   # light-blue → neon-green boundary

        from models import EmailReview
        pending_reviews = (
            EmailReview.query
            .filter_by(review_status="pending")
            .order_by(EmailReview.created_at.desc())
            .limit(30)
            .all()
        )

        return render_template(
            "dashboard.html",
            metrics=metrics,
            followups=followups,
            pipeline=pipeline,
            recent_jobs=recent_jobs,
            scan_logs=scan_logs,
            score_t1=score_t1,
            score_t2=score_t2,
            pending_reviews=pending_reviews,
        )

    # ------------------------------------------------------------------
    # Jobs
    # ------------------------------------------------------------------

    @app.route("/jobs")
    def jobs():
        category = request.args.get("category", "")
        source = request.args.get("source", "")
        status = request.args.get("status", "")
        min_score = request.args.get("min_score", "0")
        sort = request.args.get("sort", "score")
        duration = request.args.get("duration", "")

        # Exclude archived jobs — they live on /jobs/archived
        query = Job.query.filter(Job.is_active == True, Job.status != "archived")

        if category:
            query = query.filter_by(job_category=category)
        if source:
            query = query.filter_by(source=source)
        if status:
            query = query.filter_by(status=status)
        if duration:
            query = query.filter_by(job_duration=duration)
        try:
            query = query.filter(Job.match_score >= int(min_score))
        except ValueError:
            pass

        if sort == "date":
            query = query.order_by(Job.found_date.desc())
        elif sort == "company":
            query = query.order_by(Job.company.asc())
        else:
            query = query.order_by(Job.match_score.desc())

        all_jobs = query.all()

        sources = [r[0] for r in db.session.query(Job.source).distinct().all()]

        # Score tier thresholds (same formula used by dashboard & notifications)
        min_score_setting = int(Setting.get("min_match_score", 50))
        tier = (100 - min_score_setting) / 3
        score_t1 = min_score_setting + tier
        score_t2 = min_score_setting + 2 * tier

        return render_template(
            "jobs.html",
            jobs=all_jobs,
            sources=sources,
            filters={"category": category, "source": source, "status": status,
                     "min_score": min_score, "sort": sort, "duration": duration},
            score_t1=score_t1,
            score_t2=score_t2,
            score_min=min_score_setting,
            is_archived_view=False,
        )

    @app.route("/jobs/archived")
    def archived_jobs():
        """Dedicated page for archived jobs — same layout as Job Listings."""
        category = request.args.get("category", "")
        source = request.args.get("source", "")
        min_score = request.args.get("min_score", "0")
        sort = request.args.get("sort", "score")
        duration = request.args.get("duration", "")

        query = Job.query.filter(Job.is_active == True, Job.status == "archived")

        if category:
            query = query.filter_by(job_category=category)
        if source:
            query = query.filter_by(source=source)
        if duration:
            query = query.filter_by(job_duration=duration)
        try:
            query = query.filter(Job.match_score >= int(min_score))
        except ValueError:
            pass

        if sort == "date":
            query = query.order_by(Job.found_date.desc())
        elif sort == "company":
            query = query.order_by(Job.company.asc())
        else:
            query = query.order_by(Job.match_score.desc())

        all_jobs = query.all()

        sources = [r[0] for r in db.session.query(Job.source).distinct().all()]

        min_score_setting = int(Setting.get("min_match_score", 50))
        tier = (100 - min_score_setting) / 3
        score_t1 = min_score_setting + tier
        score_t2 = min_score_setting + 2 * tier

        return render_template(
            "jobs.html",
            jobs=all_jobs,
            sources=sources,
            filters={"category": category, "source": source, "status": "",
                     "min_score": min_score, "sort": sort, "duration": duration},
            score_t1=score_t1,
            score_t2=score_t2,
            score_min=min_score_setting,
            is_archived_view=True,
        )

    @app.route("/jobs/<int:job_id>")
    def job_detail(job_id):
        job = Job.query.get_or_404(job_id)
        return render_template("job_detail.html", job=job)

    @app.route("/jobs/<int:job_id>/save", methods=["POST"])
    def save_job(job_id):
        job = Job.query.get_or_404(job_id)
        job.status = "saved"
        db.session.commit()
        flash(f'"{job.title}" saved.', "success")
        return redirect(request.referrer or url_for("jobs"))

    @app.route("/jobs/<int:job_id>/archive", methods=["POST"])
    def archive_job(job_id):
        job = Job.query.get_or_404(job_id)
        job.status = "archived"
        db.session.commit()
        flash(f'"{job.title}" archived.', "info")
        return redirect(request.referrer or url_for("jobs"))

    @app.route("/jobs/<int:job_id>/unarchive", methods=["POST"])
    def unarchive_job(job_id):
        job = Job.query.get_or_404(job_id)
        job.status = "new"
        db.session.commit()
        flash(f'"{job.title}" moved back to Job Listings.', "info")
        return redirect(request.referrer or url_for("archived_jobs"))

    @app.route("/jobs/<int:job_id>/apply", methods=["POST"])
    def start_application(job_id):
        job = Job.query.get_or_404(job_id)
        if not job.application:
            cfg = _settings_as_config()
            from cover_letter import generate
            use_ai = Setting.get("use_ai_cover_letter", "false") == "true"
            letter = generate(job, cfg, use_ai=use_ai)
            follow_days = int(Setting.get("follow_up_days", "7"))

            app_obj = Application(
                job_id=job.id,
                status="draft",
                cover_letter=letter,
                next_follow_up=datetime.utcnow() + timedelta(days=follow_days),
            )
            db.session.add(app_obj)
            job.status = "applied"
        else:
            job.status = "applied"

        db.session.commit()
        flash("Application draft created.", "success")
        return redirect(url_for("application_detail", app_id=job.application.id))

    @app.route("/jobs/<int:job_id>/quick-apply", methods=["POST"])
    def quick_apply(job_id):
        """User applied directly on the job website — mark as Submitted."""
        job = Job.query.get_or_404(job_id)
        if not job.application:
            app_obj = Application(
                job_id=job.id,
                status="submitted",
                applied_date=datetime.now(),
            )
            db.session.add(app_obj)
        else:
            app_obj = job.application
            app_obj.status = "submitted"
            if not app_obj.applied_date:
                app_obj.applied_date = datetime.now()
        job.status = "applied"
        db.session.commit()
        return jsonify({
            "ok": True,
            "app_id": job.application.id,
            "redirect": url_for("application_detail", app_id=job.application.id),
        })

    @app.route("/jobs/<int:job_id>/not-applied", methods=["POST"])
    def not_applied(job_id):
        """User visited the job website but chose not to apply — record the decision."""
        job = Job.query.get_or_404(job_id)
        data = request.get_json(silent=True) or {}
        reason = (data.get("reason") or "").strip()
        note = ("Did not apply: " + reason) if reason else "Did not apply"
        if not job.application:
            app_obj = Application(
                job_id=job.id,
                status="draft",
                notes=note,
            )
            db.session.add(app_obj)
        else:
            app_obj = job.application
            app_obj.notes = ((app_obj.notes or "") + "\n" + note).strip()
            # Only downgrade status if not already in an active pipeline stage
            if app_obj.status in ("draft",):
                app_obj.status = "draft"
        # Mark the job as viewed (seen but not pursued) unless already applied
        if job.status not in ("applied",):
            job.status = "viewed"
        db.session.commit()
        return jsonify({"ok": True})

    @app.route("/jobs/<int:job_id>/cover-letter")
    def get_cover_letter(job_id):
        job = Job.query.get_or_404(job_id)
        cfg = _settings_as_config()
        from cover_letter import generate
        use_ai = Setting.get("use_ai_cover_letter", "false") == "true"
        letter = generate(job, cfg, use_ai=use_ai)
        return jsonify({"cover_letter": letter})

    @app.route("/jobs/<int:job_id>/send-application", methods=["POST"])
    def send_application_email(job_id):
        job = Job.query.get_or_404(job_id)
        cfg = _settings_as_config()
        cover_letter_text = request.form.get("cover_letter", "")
        if not cover_letter_text and job.application:
            cover_letter_text = job.application.cover_letter or ""

        from email_service import send_application
        ok = send_application(cfg, job, cover_letter_text)
        if ok:
            flash("Application email sent!", "success")
        else:
            flash("Email send failed. Check your SMTP settings.", "danger")
        return redirect(request.referrer or url_for("job_detail", job_id=job_id))

    # ------------------------------------------------------------------
    # Applications
    # ------------------------------------------------------------------

    @app.route("/applications")
    def applications():
        status_filter = request.args.get("status", "")
        query = Application.query.join(Job)
        if status_filter:
            query = query.filter(Application.status == status_filter)
        apps = query.order_by(Application.updated_at.desc()).all()
        return render_template("applications.html", apps=apps, status_filter=status_filter)

    @app.route("/applications/<int:app_id>")
    def application_detail(app_id):
        from calendar_service import interview_web_links, followup_web_links
        app_obj = Application.query.get_or_404(app_id)
        resume_default  = Setting.get("resume_path") or ""
        interview_links = interview_web_links(app_obj) if app_obj.interview_date else None
        followup_links  = followup_web_links(app_obj)
        return render_template("application_detail.html", app=app_obj,
                               resume_default=resume_default,
                               interview_links=interview_links,
                               followup_links=followup_links)

    @app.route("/api/applications/<int:app_id>/autofill")
    def api_application_autofill(app_id):
        import re, os
        app_obj = Application.query.get_or_404(app_id)
        job = app_obj.job
        result = {}

        # Resume version — derive from the settings resume_path filename
        resume_path = Setting.get("resume_path") or ""
        if resume_path:
            basename = os.path.splitext(os.path.basename(resume_path))[0]
            result["resume_version"] = basename or resume_path

        desc = job.description or ""

        # Contact email — regex scan of description
        email_m = re.search(r'[\w.+\-]+@[\w.\-]+\.\w{2,}', desc)
        if email_m:
            result["contact_email"] = email_m.group(0)

        # Contact name — common recruiter/manager phrases
        name_patterns = [
            r'(?:contact|recruiter|hiring manager|reach out to|apply to|managed by)[:\s]+([A-Z][a-z]+(?:\s[A-Z][a-z]+)+)',
            r'([A-Z][a-z]+(?:\s[A-Z][a-z]+)+)\s+is\s+(?:recruiting|hiring)',
            r'Questions\??[:\s]+([A-Z][a-z]+(?:\s[A-Z][a-z]+)+)',
        ]
        for pat in name_patterns:
            m = re.search(pat, desc)
            if m:
                result["contact_name"] = m.group(1).strip()
                break

        # Salary — parse upper bound from salary_range string
        if job.salary_range:
            nums = re.findall(r'\d[\d,]*', job.salary_range)
            nums = [int(n.replace(',', '')) for n in nums]
            if nums:
                result["salary_offered"] = max(nums)

        # Follow-up date — today + follow_up_days setting
        follow_up_days = int(Setting.get("follow_up_days") or 7)
        result["next_follow_up"] = (
            datetime.utcnow() + timedelta(days=follow_up_days)
        ).strftime("%Y-%m-%d")

        # Notes pre-fill template
        parts = [f"Position: {job.title} at {job.company}"]
        loc = job.location or ""
        if loc and job.is_remote:
            parts.append(f"Location: {loc} (Remote)")
        elif loc:
            parts.append(f"Location: {loc}")
        else:
            parts.append("Location: Remote")
        if job.salary_range:
            parts.append(f"Posted salary: {job.salary_range}")
        skills = job.get_tags()[:8]
        if skills:
            parts.append(f"Key skills: {', '.join(skills)}")
        parts.extend(["", "Notes:"])
        result["notes"] = "\n".join(parts)

        return jsonify(result)

    @app.route("/applications/<int:app_id>/update", methods=["POST"])
    def update_application(app_id):
        app_obj = Application.query.get_or_404(app_id)
        data = request.form

        if "status" in data:
            app_obj.status = data["status"]
            # Stamp applied_date the first time the application leaves draft status.
            # Use local time (datetime.now) so .date() maps to the correct local calendar day.
            if data["status"] != "draft" and not app_obj.applied_date:
                app_obj.applied_date = datetime.now()

        if "applied_date" in data:
            applied_date_str = data["applied_date"].strip()
            if applied_date_str:
                try:
                    app_obj.applied_date = datetime.strptime(applied_date_str, "%Y-%m-%d")
                except ValueError:
                    pass
            else:
                app_obj.applied_date = None

        if "cover_letter" in data:
            app_obj.cover_letter = data["cover_letter"]
        if "notes" in data:
            app_obj.notes = data["notes"]
        if "contact_name" in data:
            app_obj.contact_name = data["contact_name"].strip() or None
        if "contact_email" in data:
            app_obj.contact_email = data["contact_email"].strip() or None
        if "resume_version" in data:
            app_obj.resume_version = data["resume_version"]
        if "interview_type" in data:
            app_obj.interview_type = data["interview_type"]
        if "salary_offered" in data and data["salary_offered"]:
            try:
                app_obj.salary_offered = int(data["salary_offered"])
            except ValueError:
                pass

        interview_date_str = data.get("interview_date", "")
        interview_date_changed = False
        if interview_date_str:
            try:
                new_interview_dt = datetime.strptime(interview_date_str, "%Y-%m-%dT%H:%M")
                if new_interview_dt != app_obj.interview_date:
                    interview_date_changed = True
                app_obj.interview_date = new_interview_dt
                app_obj.status = "interview"
                # Create interview notification
                notif = Notification(
                    type="interview",
                    title=f"Interview scheduled: {app_obj.job.title}",
                    message=f"At {app_obj.job.company} on {app_obj.interview_date.strftime('%B %d, %Y %I:%M %p')}",
                    application_id=app_obj.id,
                    job_id=app_obj.job_id,
                )
                db.session.add(notif)
            except ValueError:
                pass

        next_followup_str = data.get("next_follow_up", "")
        followup_date_changed = False
        if next_followup_str:
            try:
                new_dt = datetime.strptime(next_followup_str, "%Y-%m-%d")
                if new_dt != app_obj.next_follow_up:
                    followup_date_changed = True
                app_obj.next_follow_up = new_dt
            except ValueError:
                pass

        app_obj.updated_at = datetime.utcnow()
        db.session.commit()

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            response = {"ok": True}
            if followup_date_changed:
                response["followup_changed"] = True
            if interview_date_changed:
                response["interview_changed"] = True
            return jsonify(response)

        flash("Application updated.", "success")
        return redirect(url_for("application_detail", app_id=app_id))

    @app.route("/applications/<int:app_id>/calendar/interview")
    def download_interview_ics(app_id):
        app_obj = Application.query.get_or_404(app_id)
        if not app_obj.interview_date:
            flash("No interview date set.", "warning")
            return redirect(url_for("application_detail", app_id=app_id))
        from calendar_service import create_interview_ics
        ics_data = create_interview_ics(app_obj)
        return send_file(
            BytesIO(ics_data),
            mimetype="text/calendar",
            as_attachment=True,
            download_name=f"interview_{app_obj.job.company.replace(' ', '_')}.ics",
        )

    @app.route("/applications/<int:app_id>/calendar/followup")
    def download_followup_ics(app_id):
        app_obj = Application.query.get_or_404(app_id)
        from calendar_service import create_followup_ics
        ics_data = create_followup_ics(app_obj)
        return send_file(
            BytesIO(ics_data),
            mimetype="text/calendar",
            as_attachment=True,
            download_name=f"followup_{app_obj.job.company.replace(' ', '_')}.ics",
        )

    @app.route("/applications/<int:app_id>/calendar/all")
    def download_combined_ics(app_id):
        """Download a single .ics containing both interview + follow-up events."""
        app_obj = Application.query.get_or_404(app_id)
        from calendar_service import create_combined_ics
        ics_data = create_combined_ics(app_obj)
        company = app_obj.job.company.replace(" ", "_")
        return send_file(
            BytesIO(ics_data),
            mimetype="text/calendar",
            as_attachment=True,
            download_name=f"job_tracker_{company}.ics",
        )

    # ------------------------------------------------------------------
    # ICS calendar feed  (subscribe once — no OAuth needed)
    # ------------------------------------------------------------------

    @app.route("/calendar/feed.ics")
    def calendar_feed():
        """Live ICS feed of all interview + follow-up events.
        Authenticated by a secret token stored in Settings.
        Subscribe to the URL in Google Calendar / Outlook / Apple Calendar.
        """
        token  = request.args.get("token", "")
        stored = Setting.get("cal_feed_token") or ""
        if not token or token != stored:
            abort(403)
        from calendar_service import generate_feed_ics
        tz_name = Setting.get("timezone") or "UTC"
        apps = (Application.query
                .join(Job)
                .filter(Job.is_active == True)  # noqa: E712
                .all())
        ics_data = generate_feed_ics(apps, tz_name=tz_name)
        resp = Response(ics_data, mimetype="text/calendar; charset=utf-8")
        resp.headers["Content-Disposition"] = "attachment; filename=job-tracker.ics"
        # Encourage calendar clients to re-poll frequently
        resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp

    @app.route("/api/calendar/regenerate-token", methods=["POST"])
    def api_calendar_regenerate_token():
        """Generate a new calendar feed token (invalidates old subscription URL)."""
        new_token = uuid.uuid4().hex
        Setting.set("cal_feed_token", new_token)
        return jsonify(ok=True, token=new_token)

    # ------------------------------------------------------------------
    # Database Management
    # ------------------------------------------------------------------

    @app.route("/api/database/reset", methods=["POST"])
    def api_database_reset():
        """Delete all jobs, applications, scan logs, and notifications; reset auto-increment counters."""
        try:
            # Count before deletion for the summary
            job_count = Job.query.count()
            app_count = Application.query.count()
            notif_count = Notification.query.count()
            scan_count = ScanLog.query.count()

            # Delete in dependency order (FK constraints)
            Notification.query.delete()
            ScanLog.query.delete()
            Application.query.delete()
            Job.query.delete()
            # Reset source job-count so the next scan is treated as a fresh start
            ScraperSource.query.update({ScraperSource.last_jobs_found: 0})
            db.session.commit()

            # Reset SQLite auto-increment counters
            for table in ("jobs", "applications", "notifications", "scan_logs"):
                try:
                    db.session.execute(
                        db.text("DELETE FROM sqlite_sequence WHERE name = :t"),
                        {"t": table},
                    )
                except Exception:
                    pass
            db.session.commit()

            return jsonify(
                ok=True,
                deleted={
                    "jobs": job_count,
                    "applications": app_count,
                    "notifications": notif_count,
                    "scan_logs": scan_count,
                },
            )
        except Exception as exc:
            db.session.rollback()
            logger.exception("Database reset failed")
            return jsonify(ok=False, error=str(exc)), 500

    @app.route("/api/database/repair", methods=["POST"])
    def api_database_repair():
        """Re-score all jobs, update categories and durations, remove orphaned notifications."""
        try:
            from filter_engine import score_job, categorise_job, detect_job_duration

            # Load current skill settings
            use_resume = Setting.get("use_resume_skills", "false") == "true"
            if use_resume:
                raw_req  = Setting.get("resume_required_skills") or ""
                raw_pref = Setting.get("resume_preferred_skills") or ""
            else:
                raw_req  = Setting.get("required_skills") or ""
                raw_pref = Setting.get("preferred_skills") or ""
            required_skills  = [s.strip() for s in raw_req.split(",")  if s.strip()]
            preferred_skills = [s.strip() for s in raw_pref.split(",") if s.strip()]

            jobs = Job.query.all()
            rescored = 0
            for job in jobs:
                tags = job.get_tags()
                score, matched = score_job(
                    job.title,
                    job.description or "",
                    tags,
                    required_skills,
                    preferred_skills,
                )
                job.match_score = score
                job.matched_skills = json.dumps(matched)
                job.job_category = categorise_job(job.title, job.description or "")
                if not job.job_duration:
                    duration = detect_job_duration(
                        job.title,
                        job.description or "",
                        tags,
                        job.salary_range or "",
                    )
                    if duration:
                        job.job_duration = duration
                rescored += 1

            # Remove orphaned notifications (application deleted but notification remains)
            orphaned = Notification.query.filter(
                Notification.application_id.isnot(None),
                ~Notification.application_id.in_(
                    db.session.query(Application.id)
                ),
            ).delete(synchronize_session=False)

            db.session.commit()
            return jsonify(ok=True, rescored=rescored, orphans_removed=orphaned)
        except Exception as exc:
            db.session.rollback()
            logger.exception("Database repair failed")
            return jsonify(ok=False, error=str(exc)), 500

    # ------------------------------------------------------------------
    # Analytics
    # ------------------------------------------------------------------

    @app.route("/analytics")
    def analytics():
        from analytics import (
            key_metrics, pipeline_counts, jobs_by_source,
            jobs_by_category, weekly_activity, score_distribution,
        )
        metrics = key_metrics()
        pipeline = pipeline_counts()
        by_source = jobs_by_source()
        by_category = jobs_by_category()
        weekly = weekly_activity()
        scores = score_distribution()
        return render_template(
            "analytics.html",
            metrics=metrics,
            pipeline_json=json.dumps(pipeline),
            by_source_json=json.dumps(by_source),
            by_category_json=json.dumps(by_category),
            weekly_json=json.dumps(weekly),
            scores_json=json.dumps(scores),
        )

    # ------------------------------------------------------------------
    # Settings
    # ------------------------------------------------------------------

    # ------------------------------------------------------------------
    # Unemployment tracker
    # ------------------------------------------------------------------

    @app.route("/unemployment")
    def unemployment():
        if Setting.get("unemployment_enabled", "false") != "true":
            return redirect(url_for("dashboard"))

        from datetime import date, timedelta

        # ── Settings ────────────────────────────────────────────────
        start_str    = Setting.get("unemployment_start_date", "")
        end_str      = Setting.get("unemployment_end_date", "")
        week_start   = int(Setting.get("unemployment_week_start", "0"))   # 0=Mon
        week_end     = int(Setting.get("unemployment_week_end",   "5"))   # 5=Sat
        unemp_url           = Setting.get("unemployment_url",   "")
        unemp_state         = Setting.get("unemployment_state", "")
        required_per_week   = int(Setting.get("unemployment_required_per_week", "3"))
        week_colors_raw     = Setting.get("unemployment_week_colors", "{}")
        try:
            week_colors = json.loads(week_colors_raw)
        except Exception:
            week_colors = {}

        # Day-name lookup for display
        day_names = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]

        start_date = end_date = None
        try:
            if start_str:
                start_date = datetime.strptime(start_str, "%Y-%m-%d").date()
        except ValueError:
            pass
        try:
            if end_str:
                end_date = datetime.strptime(end_str, "%Y-%m-%d").date()
            elif start_date:
                # Default: 1 year from start date
                end_date = start_date.replace(year=start_date.year + 1) - timedelta(days=1)
        except (ValueError, OverflowError):
            pass

        # ── Weekly claim periods ─────────────────────────────────────
        # Load all non-draft applications once; filter in Python per period.
        # all_apps includes withdrawn (for display in the accordion list).
        # counted_apps excludes withdrawn (for counts, met-check, and calendar).
        all_apps = (
            Application.query
            .filter(Application.status != "draft")
            .order_by(Application.created_at)
            .all()
        )
        counted_apps = [a for a in all_apps if a.status != "withdrawn"]

        def _app_date(a):
            """Return the effective date of an application (applied_date fallback created_at)."""
            d = a.applied_date or a.created_at
            return d.date() if d else None

        periods = []
        if start_date and end_date:
            today = date.today()
            # First period starts on the first occurrence of week_start on/after start_date
            days_offset = (week_start - start_date.weekday()) % 7
            current = start_date + timedelta(days=days_offset)
            days_to_end = (week_end - week_start) % 7  # length of period minus 1 day
            period_num = 1

            while current <= min(end_date, today):
                p_end = min(current + timedelta(days=days_to_end), end_date)
                # All non-draft apps in range (shown in accordion; withdrawn greyed out)
                apps_in    = [a for a in all_apps
                              if _app_date(a) and current <= _app_date(a) <= p_end]
                # Only countable (non-withdrawn) apps — drives met/completion/badge
                counted_in = [a for a in counted_apps
                              if _app_date(a) and current <= _app_date(a) <= p_end]
                app_count  = len(counted_in)
                is_current = current <= today <= p_end
                is_past    = p_end < today
                met        = app_count >= required_per_week
                completion = min(app_count / max(required_per_week, 1), 1.0)
                periods.append({
                    "num":            period_num,
                    "start":          current,
                    "end":            p_end,
                    "applications":   apps_in,
                    "counted_count":  app_count,
                    "is_current":     is_current,
                    "is_past":        is_past,
                    "met":            met,
                    "completion":     round(completion, 4),
                    "color_override": week_colors.get(str(period_num), ""),
                })
                current += timedelta(days=7)
                period_num += 1

        total_applied = sum(p["counted_count"] for p in periods)
        today = date.today()
        upload_date = (end_date + timedelta(days=1)) if end_date else None

        # Build calendar data as plain Python objects.
        # The template renders these with the tojson filter so Jinja2
        # auto-escaping never corrupts the JSON (script tags do NOT decode
        # HTML entities, so escaped quotes would break JSON.parse).
        # Use counted_apps (non-withdrawn) so withdrawn apps don't appear as
        # calendar dots.  Build from all counted apps (not period-assigned) to
        # prevent dots from "moving" when a UTC timestamp crosses a period boundary.
        _cal_applied: dict = {}
        for a in counted_apps:
            d = _app_date(a)
            if d:
                key = d.isoformat()
                _cal_applied[key] = _cal_applied.get(key, 0) + 1
        cal_applied_data = _cal_applied
        cal_periods_data = [
            {
                "num":           p["num"],
                "start":         p["start"].isoformat(),
                "end":           p["end"].isoformat(),
                "isCurrent":     p["is_current"],
                "isPast":        p["is_past"],
                "met":           p["met"],
                "completion":    p["completion"],
                "colorOverride": p["color_override"],
            }
            for p in periods
        ]

        return render_template(
            "unemployment.html",
            start_date=start_date,
            end_date=end_date,
            unemp_url=unemp_url,
            unemp_state=unemp_state,
            periods=periods,
            total_applied=total_applied,
            required_per_week=required_per_week,
            cal_periods_data=cal_periods_data,
            cal_applied_data=cal_applied_data,
            upload_date=upload_date,
            today=today,
            day_names=day_names,
            week_start=week_start,
            week_end=week_end,
        )

    @app.route("/unemployment/week/jobs")
    def unemployment_week_jobs():
        """Open a new-tab view of all jobs applied for in a specific claim week."""
        start_str = request.args.get("start", "")
        end_str   = request.args.get("end",   "")
        week_num  = request.args.get("week",  "")
        try:
            from datetime import date as _date
            w_start = datetime.strptime(start_str, "%Y-%m-%d").date()
            w_end   = datetime.strptime(end_str,   "%Y-%m-%d").date()
        except ValueError:
            abort(400, "Invalid date range.")

        def _app_date(a):
            d = a.applied_date or a.created_at
            return d.date() if d else None

        apps = (
            Application.query
            .filter(Application.status != "draft")
            .filter(Application.status != "withdrawn")
            .order_by(Application.applied_date, Application.created_at)
            .all()
        )
        week_apps = [a for a in apps if _app_date(a) and w_start <= _app_date(a) <= w_end]
        return render_template(
            "week_jobs.html",
            apps=week_apps,
            week_start=w_start,
            week_end=w_end,
            week_num=week_num,
            now=datetime.now(),
        )

    @app.route("/unemployment/week/export")
    def unemployment_week_export():
        """Download jobs applied for in a specific claim week as CSV or Excel."""
        start_str = request.args.get("start", "")
        end_str   = request.args.get("end",   "")
        week_num  = request.args.get("week",  "")
        filename  = (request.args.get("filename", "") or "week_applications").strip()
        fmt       = request.args.get("format", "csv").lower()

        try:
            w_start = datetime.strptime(start_str, "%Y-%m-%d").date()
            w_end   = datetime.strptime(end_str,   "%Y-%m-%d").date()
        except ValueError:
            abort(400, "Invalid date range.")

        def _app_date(a):
            d = a.applied_date or a.created_at
            return d.date() if d else None

        apps = (
            Application.query
            .filter(Application.status != "draft")
            .filter(Application.status != "withdrawn")
            .order_by(Application.applied_date, Application.created_at)
            .all()
        )
        week_apps = [a for a in apps if _app_date(a) and w_start <= _app_date(a) <= w_end]

        col_headers = [
            "Company", "Job Title", "Date Applied", "Company Address",
            "Contact", "Contact Email", "Status", "Job URL",
        ]

        def _row(a):
            d = a.applied_date or a.created_at
            return [
                (a.job.company         if a.job else "") or "",
                (a.job.title           if a.job else "") or "",
                d.strftime("%Y-%m-%d") if d else "",
                (a.job.company_address if a.job else "") or "",
                a.contact_name  or "",
                a.contact_email or "",
                a.status_label(),
                (a.job.url if a.job else "") or "",
            ]

        if fmt == "excel":
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"Week {week_num} Apps".strip() if week_num else "Applications"
                ws.append(col_headers)
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1E3A5F")
                for a in week_apps:
                    ws.append(_row(a))
                for col in ws.columns:
                    width = max((len(str(c.value or "")) for c in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(width + 4, 60)
                buf = BytesIO()
                wb.save(buf)
                buf.seek(0)
                safe = filename if filename.lower().endswith(".xlsx") else filename + ".xlsx"
                return send_file(
                    buf,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    as_attachment=True,
                    download_name=safe,
                )
            except ImportError:
                fmt = "csv"  # openpyxl not installed — fall back to CSV

        # CSV
        import csv as _csv
        from io import StringIO as _SIO
        buf = _SIO()
        writer = _csv.writer(buf)
        writer.writerow(col_headers)
        for a in week_apps:
            writer.writerow(_row(a))
        safe = filename if filename.lower().endswith(".csv") else filename + ".csv"
        return Response(
            buf.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{safe}"'},
        )

    @app.route("/unemployment/week-color", methods=["POST"])
    def unemp_week_color():
        """AJAX: save or clear a per-week manual color override."""
        week_num = request.form.get("week_num", "").strip()
        color    = request.form.get("color", "").strip()
        if not week_num:
            return ("", 400)
        raw = Setting.get("unemployment_week_colors", "{}")
        try:
            colors = json.loads(raw)
        except Exception:
            colors = {}
        if color:
            colors[week_num] = color
        elif week_num in colors:
            del colors[week_num]
        Setting.set("unemployment_week_colors", json.dumps(colors))
        return ("", 204)

    # ------------------------------------------------------------------
    # Settings
    # ------------------------------------------------------------------

    @app.route("/settings", methods=["GET", "POST"])
    def settings():
        if request.method == "POST":
            updatable = [
                "applicant_name", "applicant_email", "applicant_phone",
                "applicant_location", "applicant_location_miles",
                "years_experience", "applicant_linkedin",
                "applicant_github", "resume_path", "min_match_score",
                "required_skills", "preferred_skills",
                "resume_required_skills", "resume_preferred_skills",
                "follow_up_days", "timezone", "use_ai_cover_letter",
                "notif_auto_archive_days", "notif_auto_delete_days",
                "daemon_name", "unemployment_state", "unemployment_url",
                "unemployment_start_date", "unemployment_end_date",
                "unemployment_week_start", "unemployment_week_end",
                "unemployment_required_per_week",
                "posted_date_range1_days", "posted_date_range1_color",
                "posted_date_range2_days", "posted_date_range2_color",
                "scan_timeout_minutes",
            ]
            for key in updatable:
                if key in request.form:
                    Setting.set(key, request.form[key])

            # Checkbox/toggle fields — must be explicitly set when unchecked
            for key in (
                "notif_auto_archive_enabled", "notif_auto_delete_enabled",
                "scan_auto_enabled", "use_resume", "show_tray_icon",
                "unemployment_enabled", "unemployment_url_lowercase",
                "unemployment_alarm_3day", "unemployment_alarm_2day", "unemployment_alarm_1day",
            ):
                Setting.set(key, "true" if key in request.form else "false")

            # Single-select radio
            Setting.set("scan_frequency", request.form.get("scan_frequency", "daily"))

            # Multi-value checkbox fields
            scan_times_list = request.form.getlist("scan_times")
            Setting.set("scan_times", ",".join(scan_times_list) if scan_times_list else "08:00")

            scan_weekdays_list = request.form.getlist("scan_weekdays")
            Setting.set("scan_weekdays", ",".join(scan_weekdays_list) if scan_weekdays_list else "0,1,2,3,4")

            scan_monthdays_list = request.form.getlist("scan_monthdays")
            Setting.set("scan_monthdays", ",".join(scan_monthdays_list) if scan_monthdays_list else "1")

            duration_filter_list = request.form.getlist("job_duration_filter")
            Setting.set("job_duration_filter", ",".join(duration_filter_list))

            # Apply new schedule immediately
            from scheduler import reschedule_scan_jobs
            reschedule_scan_jobs(app)

            flash("Settings saved.", "success")
            return redirect(url_for("settings"))

        rows = Setting.query.order_by(Setting.key).all()
        settings_dict = {r.key: r.value for r in rows}
        return render_template("settings.html", settings=settings_dict)

    # ------------------------------------------------------------------
    # Manual scan trigger
    # ------------------------------------------------------------------

    @app.route("/scan", methods=["POST"])
    def manual_scan():
        from scheduler import run_job_scan
        t = threading.Thread(target=run_job_scan, args=[app], daemon=True)
        t.start()
        flash("Scan started in the background. Refresh in a moment.", "info")
        return redirect(request.referrer or url_for("dashboard"))

    @app.route("/api/scan/start", methods=["POST"])
    def api_scan_start():
        from scheduler import run_job_scan, get_scan_progress
        if get_scan_progress()["running"]:
            return jsonify({"ok": False, "error": "A scan is already running"})
        t = threading.Thread(target=run_job_scan, args=[app], daemon=True)
        t.start()
        return jsonify({"ok": True})

    @app.route("/api/scan/progress")
    def api_scan_progress():
        from scheduler import get_scan_progress
        return jsonify(get_scan_progress())

    # ------------------------------------------------------------------
    # Scraper Management
    # ------------------------------------------------------------------

    @app.route("/scrapers")
    def scrapers():
        from scrapers import registry
        sources = ScraperSource.query.order_by(
            ScraperSource.is_builtin.desc(), ScraperSource.display_name
        ).all()
        return render_template("scrapers.html", sources=sources)

    @app.route("/scrapers/<name>/toggle", methods=["POST"])
    def toggle_scraper(name):
        row = ScraperSource.query.filter_by(name=name).first_or_404()
        row.is_enabled = not row.is_enabled
        db.session.commit()
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify(ok=True, is_enabled=row.is_enabled)
        state = "enabled" if row.is_enabled else "disabled"
        flash(f'"{row.display_name}" {state}.', "success")
        return redirect(url_for("scrapers"))

    @app.route("/scrapers/add", methods=["POST"])
    def add_scraper():
        name = request.form.get("name", "").strip().lower().replace(" ", "_")
        display_name = request.form.get("display_name", "").strip()
        source_type = request.form.get("source_type", "rss")
        url_template = request.form.get("url_template", "").strip()
        description = request.form.get("description", "").strip()
        search_terms_raw = request.form.get("search_terms", "").strip()

        _is_xhr = request.headers.get("X-Requested-With") == "XMLHttpRequest"
        if not name or not url_template:
            if _is_xhr:
                return jsonify(ok=False, error="Name and URL template are required.")
            flash("Name and URL template are required.", "danger")
            return redirect(url_for("scrapers"))

        if ScraperSource.query.filter_by(name=name).first():
            if _is_xhr:
                return jsonify(ok=False, error=f'A source named "{name}" already exists.')
            flash(f'A source named "{name}" already exists.', "warning")
            return redirect(url_for("scrapers"))

        # Parse search terms: one per line or comma-separated
        if search_terms_raw:
            if "\n" in search_terms_raw:
                terms = [t.strip() for t in search_terms_raw.splitlines() if t.strip()]
            else:
                terms = [t.strip() for t in search_terms_raw.split(",") if t.strip()]
        else:
            terms = []

        row = ScraperSource(
            name=name,
            display_name=display_name or name,
            description=description,
            source_type=source_type,
            url_template=url_template,
            is_builtin=False,
            is_enabled=True,
            search_terms=json.dumps(terms),
        )
        db.session.add(row)
        db.session.commit()
        if _is_xhr:
            return jsonify(ok=True, name=name)
        flash(f'Scraper "{display_name or name}" added.', "success")
        return redirect(url_for("scrapers"))

    @app.route("/scrapers/<name>/edit", methods=["POST"])
    def edit_scraper(name):
        row = ScraperSource.query.filter_by(name=name).first_or_404()
        row.display_name  = request.form.get("display_name", row.display_name).strip()
        row.description   = request.form.get("description", row.description or "").strip()
        row.url_template  = request.form.get("url_template", row.url_template or "").strip()

        search_terms_raw = request.form.get("search_terms", "").strip()
        if search_terms_raw:
            if "\n" in search_terms_raw:
                terms = [t.strip() for t in search_terms_raw.splitlines() if t.strip()]
            else:
                terms = [t.strip() for t in search_terms_raw.split(",") if t.strip()]
            row.search_terms = json.dumps(terms)

        db.session.commit()
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify(ok=True)
        flash(f'"{row.display_name}" updated.', "success")
        return redirect(url_for("scrapers"))

    @app.route("/scrapers/<name>/delete", methods=["POST"])
    def delete_scraper(name):
        row = ScraperSource.query.filter_by(name=name).first_or_404()
        _is_xhr = request.headers.get("X-Requested-With") == "XMLHttpRequest"
        if row.is_builtin:
            if _is_xhr:
                return jsonify(ok=False, error="Built-in scrapers cannot be deleted (you can disable them).")
            flash("Built-in scrapers cannot be deleted (you can disable them).", "warning")
            return redirect(url_for("scrapers"))
        db.session.delete(row)
        db.session.commit()
        if _is_xhr:
            return jsonify(ok=True)
        flash(f'"{row.display_name}" deleted.', "info")
        return redirect(url_for("scrapers"))

    @app.route("/scrapers/<name>/run", methods=["POST"])
    def run_single_scraper(name):
        """Run one specific scraper immediately in a background thread."""
        row = ScraperSource.query.filter_by(name=name).first_or_404()

        def _run():
            from scrapers import registry, lookup_company_info
            from filter_engine import score_job, categorise_job, is_relevant, location_passes, detect_job_duration
            from config import Config

            cfg = Config()
            scraper = registry._resolve(row)
            if not scraper:
                return

            with app.app_context():
                from models import Job, Notification
                from extensions import db

                applicant_state = Setting.get("applicant_location", "Remote")
                applicant_miles = int(Setting.get("applicant_location_miles", "25"))
                min_match_score = int(Setting.get("min_match_score", str(cfg.MIN_MATCH_SCORE)))

                if Setting.get("use_resume", "false") == "true":
                    _required  = [s.strip() for s in Setting.get("resume_required_skills", "").split(",") if s.strip()]
                    _preferred = [s.strip() for s in Setting.get("resume_preferred_skills", "").split(",") if s.strip()]
                else:
                    _req_str  = Setting.get("required_skills",  ",".join(cfg.REQUIRED_SKILLS))
                    _pref_str = Setting.get("preferred_skills", ",".join(cfg.PREFERRED_SKILLS))
                    _required  = [s.strip() for s in _req_str.split(",")  if s.strip()]
                    _preferred = [s.strip() for s in _pref_str.split(",") if s.strip()]

                terms = row.get_search_terms() or None
                try:
                    raw_jobs = scraper.fetch(search_terms=terms)
                    row.last_run = datetime.utcnow()
                    row.last_jobs_found = 0
                    new_count = 0

                    for raw in raw_jobs:
                        ext_id = raw.get("external_id", "")
                        if not ext_id or Job.query.filter_by(external_id=ext_id).first():
                            continue

                        score, matched = score_job(
                            raw.get("title", ""), raw.get("description", ""),
                            raw.get("tags", []),
                            required_skills=_required,
                            preferred_skills=_preferred,
                        )
                        if not is_relevant(score, min_match_score):
                            continue

                        if not location_passes(raw.get("location", ""), applicant_state, applicant_miles):
                            continue

                        _title = raw.get("title", "")
                        _desc  = raw.get("description", "")
                        _tags  = raw.get("tags", [])
                        _co_info = lookup_company_info(raw.get("company", ""))
                        job = Job(
                            external_id=ext_id,
                            title=_title,
                            company=raw.get("company", ""),
                            location=raw.get("location", "Remote"),
                            is_remote=True,
                            job_duration=detect_job_duration(_title, _desc, _tags) or None,
                            job_category=categorise_job(_title, _desc),
                            description=_desc,
                            tags=json.dumps(_tags),
                            salary_range=raw.get("salary_range", ""),
                            url=raw.get("url", ""),
                            source=raw.get("source", name),
                            posted_date=raw.get("posted_date"),
                            found_date=datetime.utcnow(),
                            match_score=score,
                            matched_skills=json.dumps(matched),
                            status="new",
                            company_address=_co_info.get("company_address", ""),
                            company_phone=_co_info.get("company_phone", ""),
                        )
                        db.session.add(job)
                        new_count += 1

                    db.session.flush()
                    row.last_status = "success"
                    row.last_jobs_found = new_count
                    row.last_error = None
                    db.session.commit()
                except Exception as exc:
                    row.last_status = "error"
                    row.last_error = str(exc)
                    db.session.commit()
                    logger.error("Single scraper '%s' failed: %s", name, exc)

        t = threading.Thread(target=_run, daemon=True)
        t.start()
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify(ok=True, message=f'"{row.display_name}" started.')
        flash(f'Scraper "{row.display_name}" started. Refresh in a moment.', "info")
        return redirect(url_for("scrapers"))

    # ------------------------------------------------------------------
    # Source Health Check & Auto-Discovery API
    # ------------------------------------------------------------------

    @app.route("/api/scrapers/health-check/start", methods=["POST"])
    def api_health_check_start():
        import threading as _th
        from source_health import run_health_check, get_health_state  # type: ignore
        state = get_health_state()
        if state["running"]:
            return jsonify({"ok": False, "error": "Health check already running"})
        _th.Thread(target=run_health_check, args=(app,), daemon=True).start()
        return jsonify({"ok": True})

    @app.route("/api/scrapers/health-check/progress")
    def api_health_check_progress():
        from source_health import get_health_state  # type: ignore
        return jsonify(get_health_state())

    @app.route("/api/scrapers/discover/start", methods=["POST"])
    def api_discover_start():
        import threading as _th
        from source_health import run_auto_discovery, get_discovery_state  # type: ignore
        state = get_discovery_state()
        if state["running"]:
            return jsonify({"ok": False, "error": "Discovery already running"})
        data = request.get_json(silent=True) or {}
        raw = data.get("urls", "")
        urls = [u.strip() for u in raw.splitlines() if u.strip()]
        if not urls:
            return jsonify({"ok": False, "error": "No URLs provided"})
        _th.Thread(target=run_auto_discovery, args=(urls,), daemon=True).start()
        return jsonify({"ok": True, "count": len(urls)})

    @app.route("/api/scrapers/discover/progress")
    def api_discover_progress():
        from source_health import get_discovery_state  # type: ignore
        return jsonify(get_discovery_state())

    @app.route("/api/scrapers/discover-boards/start", methods=["POST"])
    def api_board_discovery_start():
        """Start skill-based job board discovery in the background."""
        import threading as _th
        from source_health import run_board_discovery, get_board_discovery_state  # type: ignore
        if get_board_discovery_state()["running"]:
            return jsonify({"ok": False, "error": "Discovery already running"})

        if Setting.get("use_resume", "false") == "true":
            req_raw  = Setting.get("resume_required_skills",  "") or ""
            pref_raw = Setting.get("resume_preferred_skills", "") or ""
        else:
            req_raw  = Setting.get("required_skills",  "") or ""
            pref_raw = Setting.get("preferred_skills", "") or ""

        all_skills = [s.strip() for s in (req_raw + "," + pref_raw).split(",") if s.strip()]
        _th.Thread(target=run_board_discovery, args=(app, all_skills), daemon=True).start()
        return jsonify({"ok": True, "skill_count": len(all_skills)})

    @app.route("/api/scrapers/discover-boards/progress")
    def api_board_discovery_progress():
        from source_health import get_board_discovery_state  # type: ignore
        return jsonify(get_board_discovery_state())

    @app.route("/api/scrapers/discover/add", methods=["POST"])
    def api_discover_add():
        """Add a discovered source to the database."""
        data = request.get_json(silent=True) or {}
        import re as _re
        raw_name = data.get("title") or data.get("url_template", "")
        slug = _re.sub(r"[^a-z0-9_]", "_",
                       raw_name.lower().split("//")[-1].split("/")[0].replace(".", "_"))[:40] or "custom"
        # Ensure unique slug
        base_slug = slug
        counter = 1
        while ScraperSource.query.filter_by(name=slug).first():
            slug = f"{base_slug}_{counter}"
            counter += 1

        row = ScraperSource(
            name=slug,
            display_name=data.get("title") or slug.replace("_", " ").title(),
            description=data.get("description", ""),
            source_type=data.get("type", "rss"),
            url_template=data.get("url_template", ""),
            is_builtin=False,
            is_enabled=True,
            search_terms=json.dumps([]),
        )
        db.session.add(row)
        db.session.commit()
        return jsonify({"ok": True, "name": slug})

    @app.route("/api/scrapers/<name>/rename", methods=["POST"])
    def api_scraper_rename(name):
        """Rename the display_name of any source (builtin or custom)."""
        data = request.get_json(silent=True) or {}
        new_display = (data.get("display_name") or "").strip()
        if not new_display:
            return jsonify(ok=False, error="Name cannot be empty")
        row = ScraperSource.query.filter_by(name=name).first_or_404()
        row.display_name = new_display
        db.session.commit()
        return jsonify(ok=True, display_name=new_display)

    # ------------------------------------------------------------------
    # Notifications
    # ------------------------------------------------------------------

    @app.route("/notifications")
    def notifications():
        from datetime import datetime as _dt
        from sqlalchemy import nullslast

        # Run auto-archive / auto-delete based on settings
        _apply_notification_cleanup()

        min_score = int(Setting.get("min_match_score", 50))
        tier      = (100 - min_score) / 3
        score_t1  = min_score + tier
        score_t2  = min_score + 2 * tier

        # Assign each notification a tier bucket (3=neon, 2=blue, 1=grey, 0=no job)
        tier_expr = db.case(
            (db.and_(Notification.job_id.isnot(None), Job.match_score >= score_t2), 3),
            (db.and_(Notification.job_id.isnot(None), Job.match_score >= score_t1), 2),
            (db.and_(Notification.job_id.isnot(None), Job.match_score.isnot(None)), 1),
            else_=0,
        )

        notifs = (
            Notification.query
            .filter(Notification.is_archived == False)  # noqa: E712
            .outerjoin(Job, Notification.job_id == Job.id)
            .order_by(
                tier_expr.desc(),                          # neon → blue → grey → no-job
                nullslast(Job.posted_date.desc()),         # newest posting first within tier
                Notification.created_at.desc(),            # tiebreaker
            )
            .limit(50)
            .all()
        )

        return render_template(
            "notifications.html",
            notifications=notifs,
            score_t1=score_t1,
            score_t2=score_t2,
            today=_dt.utcnow().date(),
        )

    @app.route("/notifications/mark-read", methods=["POST"])
    def mark_notifications_read():
        Notification.query.update({"is_read": True})
        db.session.commit()
        return redirect(url_for("notifications"))

    # ------------------------------------------------------------------
    # API endpoints (JSON)
    # ------------------------------------------------------------------

    @app.route("/api/jobs")
    def api_jobs():
        jobs = (
            Job.query
            .filter(Job.is_active == True)
            .order_by(Job.match_score.desc())
            .limit(100)
            .all()
        )
        return jsonify([{
            "id": j.id, "title": j.title, "company": j.company,
            "score": j.match_score, "status": j.status,
            "category": j.job_category, "url": j.url,
            "found": j.found_date.isoformat() if j.found_date else None,
        } for j in jobs])

    # ------------------------------------------------------------------
    # Help / Documentation page
    # ------------------------------------------------------------------

    @app.route("/help")
    def help_page():
        return render_template("help.html")

    @app.route("/support")
    def support_page():
        applicant_name  = Setting.get("applicant_name", "")
        applicant_email = Setting.get("applicant_email", "")
        return render_template("support.html",
                               applicant_name=applicant_name,
                               applicant_email=applicant_email)

    # ------------------------------------------------------------------
    # Server Log page
    # ------------------------------------------------------------------

    @app.route("/server")
    def server_page():
        import platform, sys
        return render_template(
            "server.html",
            py_version=sys.version.split()[0],
            platform_name=platform.system(),
            platform_release=platform.release(),
        )

    @app.route("/api/server/log")
    def api_server_log():
        from server_log import get_log_entries, get_capture_mode
        since = float(request.args.get("since", 0))
        return jsonify({
            "entries":      get_log_entries(since),
            "capture_mode": get_capture_mode(),
        })

    @app.route("/api/server/log/capture", methods=["POST"])
    def api_server_log_capture():
        from server_log import set_capture_mode, get_capture_mode
        data = request.get_json(silent=True) or {}
        mode = data.get("mode", "")
        if mode in ("off", "warn", "all"):
            set_capture_mode(mode)
        return jsonify({"capture_mode": get_capture_mode()})

    @app.route("/api/server/log/clear", methods=["POST"])
    def api_server_log_clear():
        from server_log import clear_log
        clear_log()
        return jsonify({"ok": True})

    # ------------------------------------------------------------------

    @app.route("/api/scan-status")
    def api_scan_status():
        log = ScanLog.query.order_by(ScanLog.scan_date.desc()).first()
        if not log:
            return jsonify({"status": "never", "last_scan": None})
        return jsonify({
            "status": log.status,
            "last_scan": log.scan_date.isoformat(),
            "jobs_found": log.jobs_found,
            "jobs_new": log.jobs_new,
            "jobs_matched": log.jobs_matched,
        })

    @app.route("/api/notifications/count")
    def api_notification_count():
        count = Notification.query.filter_by(is_read=False, is_archived=False).count()
        return jsonify({"unread": count})

    @app.route("/api/notifications/<int:notif_id>/acknowledge", methods=["POST"])
    def api_notification_acknowledge(notif_id):
        notif = Notification.query.get_or_404(notif_id)
        data = request.get_json(silent=True) or {}
        notif.is_acknowledged = data.get("acknowledged", True)
        db.session.commit()
        return jsonify({"ok": True, "is_acknowledged": notif.is_acknowledged})

    @app.route("/api/skills-taxonomy")
    def api_skills_taxonomy():
        """Return the merged skills taxonomy (static + scan-discovered) for the Skills Picker modal."""
        from skills_taxonomy import get_taxonomy
        taxonomy = dict(get_taxonomy())

        # Merge in skills discovered during job scans
        try:
            discovered = json.loads(Setting.get("discovered_skills", "{}") or "{}")
            non_empty = {k: v for k, v in discovered.items() if v}
            if non_empty:
                taxonomy["Discovered Skills"] = non_empty
        except Exception:
            pass

        return jsonify(taxonomy)

    @app.route("/api/browse-file")
    def api_browse_file():
        """Open a native OS file picker and return the chosen path (local app only)."""
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.wm_attributes("-topmost", True)
            path = filedialog.askopenfilename(
                title="Select Resume File",
                filetypes=[
                    ("PDF files", "*.pdf"),
                    ("Word documents", "*.docx *.doc"),
                    ("All files", "*.*"),
                ],
            )
            root.destroy()
            return jsonify({"path": path or ""})
        except Exception as exc:
            logger.warning("browse-file dialog failed: %s", exc)
            return jsonify({"path": "", "error": str(exc)})

    @app.route("/api/parse-resume")
    def api_parse_resume():
        """Parse the configured resume file and return skills matched against the taxonomy."""
        resume_path = Setting.get("resume_path", "")
        if not resume_path:
            return jsonify({"error": "No resume path configured in Settings.", "skills": []})
        try:
            from resume_parser import extract_skills_from_resume, pdf_library_available
            # Check whether we'll need to auto-install a PDF library
            is_pdf = not resume_path.lower().endswith((".docx", ".doc"))
            had_to_install = is_pdf and not pdf_library_available()
            skills = extract_skills_from_resume(resume_path)
            return jsonify({
                "skills": skills,
                "count": len(skills),
                "installed_library": had_to_install,
            })
        except Exception as exc:
            logger.warning("Resume parse failed: %s", exc)
            return jsonify({"error": str(exc), "skills": []})

    # ------------------------------------------------------------------ #
    #  Scheduler Daemon management endpoints                               #
    # ------------------------------------------------------------------ #

    import os as _os, signal as _signal, subprocess as _subprocess, sys as _sys

    _DAEMON_SCRIPT = _os.path.join(_os.path.dirname(__file__), "scheduler_daemon.py")
    _PID_FILE      = _os.path.join(_os.path.dirname(__file__), "scheduler_daemon.pid")

    def _build_version_info_blob(description: str) -> bytes:
        """
        Build a valid VS_VERSIONINFO binary resource blob.
        FileDescription is set to *description*; all version numbers are 0.
        Injecting this via UpdateResourceW makes Task Manager Processes tab
        show *description* instead of the original 'Python 3.x' string.
        """
        import struct as _st

        def _wstr(s: str) -> bytes:
            return (s + "\0").encode("utf-16-le")

        def _block(key: str, value_bytes: bytes, wtype: int, children: bytes) -> bytes:
            """
            Build one VERSION_INFO block:
              WORD wLength, WORD wValueLength, WORD wType
              WCHAR szKey[] (null-terminated)
              WORD  Padding1 (align header+key to DWORD)
              Bytes value
              WORD  Padding2 (align to DWORD before children)
              Children
            wLength = total bytes including all padding (no trailing sibling pad).
            """
            k     = _wstr(key)
            lenAB = 6 + len(k)
            lenC  = (4 - lenAB % 4) % 4
            lenD  = len(value_bytes)
            lenE  = (4 - (lenAB + lenC + lenD) % 4) % 4
            w_len = lenAB + lenC + lenD + lenE + len(children)
            w_vl  = len(value_bytes) // 2 if wtype == 1 else len(value_bytes)
            return (
                _st.pack("<HHH", w_len, w_vl, wtype)
                + k
                + b"\x00" * lenC
                + value_bytes
                + b"\x00" * lenE
                + children
            )

        def _p4(b: bytes) -> bytes:
            r = len(b) % 4
            return b + b"\x00" * ((4 - r) % 4)

        # String "FileDescription" = description
        desc_block = _block("FileDescription", _wstr(description), 1, b"")
        # StringTable "040904b0"
        st_block   = _block("040904b0",      b"", 1, _p4(desc_block))
        # StringFileInfo
        sfi_block  = _block("StringFileInfo", b"", 1, _p4(st_block))
        # Var "Translation" (lang=English US, codepage=Unicode)
        var_block  = _block("Translation", _st.pack("<HH", 0x0409, 0x04B0), 0, b"")
        # VarFileInfo
        vfi_block  = _block("VarFileInfo",  b"", 1, _p4(var_block))
        # VS_FIXEDFILEINFO (52 bytes, all versions 0)
        ffi = _st.pack(
            "<13I",
            0xFEEF04BD, 0x00010000,   # signature, struct version
            0, 0, 0, 0,               # file / product version
            0x3F, 0,                  # flags mask, flags
            0x00040004, 0x00000001,   # fileOS (NT|Win32), fileType (APP)
            0, 0, 0,                  # fileSubtype, fileDate MS/LS
        )
        # VS_VERSIONINFO (root)
        root_children = _p4(sfi_block) + _p4(vfi_block)
        return _block("VS_VERSION_INFO", ffi, 0, root_children)

    def _set_exe_description(exe_path: str, description: str) -> None:
        """
        Write a VS_VERSIONINFO resource into *exe_path* that sets
        FileDescription to *description*.  After this, Windows Task Manager
        Processes tab shows *description* instead of 'Python 3.x'.

        Uses EnumResourceLanguagesW (with correct LONG_PTR size) to discover
        and delete existing RT_VERSION entries before writing the new one, so
        only one version-info entry exists.  No-op when the exe is locked
        (i.e. daemon already running — try again after stopping it).
        """
        if _sys.platform != "win32":
            return
        try:
            import ctypes as _ct
            # use_last_error=True lets us call _ct.get_last_error() reliably
            _k32 = _ct.WinDLL("kernel32", use_last_error=True)

            _k32.LoadLibraryExW.restype  = _ct.c_void_p
            _k32.LoadLibraryExW.argtypes = [_ct.c_wchar_p, _ct.c_void_p, _ct.c_ulong]
            _k32.FreeLibrary.restype     = _ct.c_bool
            _k32.FreeLibrary.argtypes    = [_ct.c_void_p]

            # LONG_PTR is 64-bit on x64 Windows → c_ssize_t
            _ELPROC = _ct.WINFUNCTYPE(
                _ct.c_bool,
                _ct.c_void_p, _ct.c_void_p, _ct.c_void_p,
                _ct.c_ushort, _ct.c_ssize_t,
            )
            _k32.EnumResourceLanguagesW.restype  = _ct.c_bool
            _k32.EnumResourceLanguagesW.argtypes = [
                _ct.c_void_p, _ct.c_void_p, _ct.c_void_p,
                _ELPROC, _ct.c_ssize_t,
            ]

            existing_langs: list = []

            @_ELPROC
            def _cb(hm, tp, nm, lang, lp):  # noqa: ANN001
                existing_langs.append(lang)
                return True

            hmod = _k32.LoadLibraryExW(exe_path, None, 0x00000002)  # LOAD_LIBRARY_AS_DATAFILE
            if hmod:
                _k32.EnumResourceLanguagesW(hmod, 16, 1, _cb, 0)
                _k32.FreeLibrary(hmod)
            if not existing_langs:
                existing_langs = [0x0000, 0x0409]

            # Build the new version info blob with description as FileDescription
            new_blob = _build_version_info_blob(description)
            buf = (_ct.c_char * len(new_blob)).from_buffer_copy(new_blob)

            _k32.BeginUpdateResourceW.restype  = _ct.c_void_p
            _k32.BeginUpdateResourceW.argtypes = [_ct.c_wchar_p, _ct.c_bool]
            _k32.UpdateResourceW.restype  = _ct.c_bool
            _k32.UpdateResourceW.argtypes = [
                _ct.c_void_p, _ct.c_void_p, _ct.c_void_p,
                _ct.c_ushort, _ct.c_void_p, _ct.c_ulong,
            ]
            _k32.EndUpdateResourceW.restype  = _ct.c_bool
            _k32.EndUpdateResourceW.argtypes = [_ct.c_void_p, _ct.c_bool]

            h = _k32.BeginUpdateResourceW(exe_path, False)
            if not h:
                logger.warning(
                    "_set_exe_description: BeginUpdateResourceW failed for '%s' "
                    "(err=%d) — exe may be in use; name not updated.",
                    exe_path, _ct.get_last_error(),
                )
                return

            # Delete all existing RT_VERSION entries
            for lang in existing_langs:
                _k32.UpdateResourceW(h, 16, 1, lang, None, 0)

            # Write the new version info under English (US)
            ok = _k32.UpdateResourceW(h, 16, 1, 0x0409, buf, len(new_blob))
            if not ok:
                logger.warning(
                    "_set_exe_description: UpdateResourceW (write) failed "
                    "(err=%d); discarding.", _ct.get_last_error(),
                )
                _k32.EndUpdateResourceW(h, True)   # discard
                return

            if _k32.EndUpdateResourceW(h, False):  # commit
                logger.info(
                    "_set_exe_description: '%s' → FileDescription='%s'",
                    exe_path, description,
                )
            else:
                logger.warning(
                    "_set_exe_description: EndUpdateResourceW commit failed (err=%d).",
                    _ct.get_last_error(),
                )
        except Exception:
            logger.exception("_set_exe_description raised for '%s'", exe_path)

    def _get_named_launcher(daemon_name: str) -> str:
        """
        Build (or refresh) a standalone named Python launcher in the app
        directory so Windows Task Manager shows *daemon_name* instead of
        'Python 3.x'.

        Strategy (works for regular installs AND Microsoft Store Python,
        which puts 0-byte App Execution Alias stubs in sys.executable's dir):

        1. Use GetModuleFileNameW to find the REAL running Python binary
           (e.g. …\\WindowsApps\\PythonSoftwareFoundation.Python.3.13_…\\pythonw.exe).
        2. Copy it to the job_tracker app directory as <DaemonName>.exe.
        3. Copy the Python DLLs needed to run outside the Python home directory.
        4. Write a pyvenv.cfg that points back to the original Python home so
           the interpreter can locate its standard library without PYTHONHOME.
        5. Inject a VS_VERSIONINFO resource with FileDescription=daemon_name
           so the Processes tab shows the custom name.

        Falls back to sys.executable on unrecoverable errors.
        """
        import re as _re, shutil as _shutil, ctypes as _ct

        safe = _re.sub(r"[^A-Za-z0-9_\-]", "_", daemon_name.strip()) or "JobTrackerDaemon"
        safe = safe[:40]

        # ── 1. Find the REAL Python binary ───────────────────────────────────
        # sys.executable is a 0-byte App Execution Alias for MS Store Python;
        # GetModuleFileNameW returns the path of the actual running image.
        _buf = _ct.create_unicode_buffer(1024)
        _ct.windll.kernel32.GetModuleFileNameW(None, _buf, 1024)
        real_exe = _buf.value or _sys.executable
        real_dir = _os.path.dirname(real_exe)

        # Prefer the windowless pythonw variant (no console); fall back to
        # the exe actually running or plain python.exe.
        _src = ""
        for _candidate in ("pythonw.exe", _os.path.basename(real_exe), "python.exe"):
            _try = _os.path.join(real_dir, _candidate)
            if _os.path.isfile(_try) and _os.path.getsize(_try) > 0:
                _src = _try
                break
        if not _src:
            logger.warning("_get_named_launcher: could not find real Python exe in %s", real_dir)
            return _sys.executable

        # ── 2. Copy exe to app directory ──────────────────────────────────────
        app_dir = _os.path.dirname(__file__)
        dest    = _os.path.join(app_dir, safe + ".exe")

        needs_copy = (
            not _os.path.exists(dest)
            or _os.path.getsize(dest) == 0
            or _os.path.getmtime(_src) > _os.path.getmtime(dest)
        )
        if needs_copy:
            try:
                _shutil.copy2(_src, dest)
                logger.info("_get_named_launcher: copied '%s' → '%s'", _src, dest)
            except Exception as _exc:
                logger.warning("_get_named_launcher: copy failed: %s", _exc)
                if not _os.path.exists(dest) or _os.path.getsize(dest) == 0:
                    return _sys.executable

            # ── 3. Copy companion DLLs ────────────────────────────────────────
            _ver = f"{_sys.version_info.major}{_sys.version_info.minor}"
            for _dll in (
                "python3.dll",
                f"python{_ver}.dll",
                "vcruntime140.dll",
                "vcruntime140_1.dll",
            ):
                _dll_src = _os.path.join(real_dir, _dll)
                _dll_dst = _os.path.join(app_dir, _dll)
                if _os.path.isfile(_dll_src) and (
                    not _os.path.exists(_dll_dst)
                    or _os.path.getmtime(_dll_src) > _os.path.getmtime(_dll_dst)
                ):
                    try:
                        _shutil.copy2(_dll_src, _dll_dst)
                    except Exception:
                        pass

            # ── 3b. Copy DLLs/ extension modules (.pyd + companion DLLs) ─────
            # The WindowsApps directory is access-restricted for processes that
            # are not Store apps, so _ctypes.pyd etc. cannot be loaded from
            # there when running as a copied exe.  We copy them to a local
            # DLLs/ subdirectory which scheduler_daemon.py inserts first into
            # sys.path via _fix_sys_path().
            _base_dlls_src = _os.path.join(real_dir, "DLLs")
            if not _os.path.isdir(_base_dlls_src):
                # Some layouts put DLLs next to python.exe
                _base_dlls_src = real_dir
            _local_dlls_dst = _os.path.join(app_dir, "DLLs")
            _os.makedirs(_local_dlls_dst, exist_ok=True)
            try:
                for _f in _os.listdir(_base_dlls_src):
                    if _f.endswith((".pyd", ".dll", ".cat")):
                        _fsrc = _os.path.join(_base_dlls_src, _f)
                        _fdst = _os.path.join(_local_dlls_dst, _f)
                        if _os.path.isfile(_fsrc) and (
                            not _os.path.exists(_fdst)
                            or _os.path.getmtime(_fsrc) > _os.path.getmtime(_fdst)
                        ):
                            try:
                                _shutil.copy2(_fsrc, _fdst)
                            except Exception:
                                pass
            except Exception:
                pass

            # ── 4. Write pyvenv.cfg so Python finds its standard library ──────
            # Python reads this file from the exe's directory at startup and
            # uses `home` to locate the Lib directory — no PYTHONHOME needed.
            _cfg_path    = _os.path.join(app_dir, "pyvenv.cfg")
            _cfg_content = f"home = {real_dir}\ninclude-system-site-packages = true\n"
            try:
                _existing = open(_cfg_path).read() if _os.path.exists(_cfg_path) else ""
                if _existing != _cfg_content:
                    with open(_cfg_path, "w") as _fh:
                        _fh.write(_cfg_content)
            except Exception as _exc:
                logger.warning("_get_named_launcher: pyvenv.cfg write failed: %s", _exc)

            # ── 5. Inject FileDescription via VS_VERSIONINFO ──────────────────
            _set_exe_description(dest, daemon_name)

        return dest

    def _daemon_status():
        """Return (running: bool, pid: int|None).

        Checks in order:
        1. PID file — fast path for daemons started via the web UI.
        2. tasklist — catches daemons auto-started at Windows login via the
           registry Run key (which bypass the PID-file write path).
        """
        # ── 1. PID file ──────────────────────────────────────────────────────
        try:
            with open(_PID_FILE) as fh:
                pid = int(fh.read().strip())
            _os.kill(pid, 0)   # raises OSError if process is gone
            return True, pid
        except FileNotFoundError:
            pass   # no PID file — fall through to tasklist check
        except (ValueError, OSError):
            try:
                _os.remove(_PID_FILE)
            except Exception:
                pass

        # ── 2. tasklist fallback (Windows only) ──────────────────────────────
        if _sys.platform == "win32":
            try:
                import re as _re
                from models import Setting
                daemon_name = Setting.get("daemon_name", "JobTrackerDaemon")
                safe     = _re.sub(r"[^A-Za-z0-9_\-]", "_", daemon_name.strip()) or "JobTrackerDaemon"
                exe_name = safe[:40] + ".exe"
                result   = _subprocess.run(
                    ["tasklist", "/FI", f"IMAGENAME eq {exe_name}", "/FO", "CSV", "/NH"],
                    capture_output=True, text=True, timeout=5,
                )
                if exe_name.lower() in result.stdout.lower():
                    import csv as _csv, io as _io
                    for row in _csv.reader(_io.StringIO(result.stdout)):
                        if len(row) >= 2:
                            try:
                                pid = int(row[1])
                                # Cache the PID so subsequent fast-path checks work
                                with open(_PID_FILE, "w") as fh:
                                    fh.write(str(pid))
                                return True, pid
                            except (ValueError, IndexError):
                                pass
                    return True, None   # process found but couldn't parse PID
            except Exception:
                pass

        return False, None

    @app.route("/api/daemon/status")
    def api_daemon_status():
        from windows_task import is_task_installed  # type: ignore
        daemon_name = Setting.get("daemon_name", "JobTrackerDaemon")
        running, pid = _daemon_status()
        return jsonify({
            "running":        running,
            "pid":            pid,
            "task_installed": is_task_installed(value_name=daemon_name),
        })

    @app.route("/api/daemon/start", methods=["POST"])
    def api_daemon_start():
        running, pid = _daemon_status()
        if running:
            return jsonify({"ok": True, "message": f"Daemon already running (PID {pid})."})
        try:
            daemon_name = Setting.get("daemon_name", "JobTrackerDaemon")
            launcher = _get_named_launcher(daemon_name)
            flags = 0
            if _sys.platform == "win32":
                flags = _subprocess.CREATE_NO_WINDOW | _subprocess.DETACHED_PROCESS
            _subprocess.Popen(
                [launcher, _DAEMON_SCRIPT],
                creationflags=flags,
                close_fds=True,
            )
            # Give it a moment to write its PID file
            import time as _time
            _time.sleep(1.5)
            running, pid = _daemon_status()
            if running:
                return jsonify({"ok": True,  "message": f"Daemon started (PID {pid})."})
            return jsonify({"ok": False, "message": "Daemon launched but PID file not yet written — try refreshing."})
        except Exception as exc:
            return jsonify({"ok": False, "message": str(exc)})

    @app.route("/api/daemon/stop", methods=["POST"])
    def api_daemon_stop():
        running, pid = _daemon_status()
        if not running:
            return jsonify({"ok": True, "message": "Daemon is not running."})

        def _force_kill(pid):
            """Kill the process and its children immediately."""
            if _sys.platform == "win32":
                # /T kills the entire process tree; /F forces immediate termination
                _subprocess.run(
                    ["taskkill", "/PID", str(pid), "/T", "/F"],
                    capture_output=True, timeout=10,
                )
            else:
                try:
                    _os.kill(pid, _signal.SIGKILL)
                except OSError:
                    pass

        def _remove_pid_file():
            try:
                _os.remove(_PID_FILE)
            except FileNotFoundError:
                pass

        try:
            import time as _time
            _force_kill(pid)
            _remove_pid_file()          # remove stale PID file immediately
            _time.sleep(0.3)            # brief settle time
            return jsonify({"ok": True, "message": f"Daemon (PID {pid}) terminated."})
        except Exception as exc:
            _remove_pid_file()          # clean up even on error
            return jsonify({"ok": False, "message": str(exc)})

    @app.route("/api/daemon/install-task", methods=["POST"])
    def api_daemon_install_task():
        from windows_task import install_task  # type: ignore
        daemon_name = Setting.get("daemon_name", "JobTrackerDaemon")
        launcher = _get_named_launcher(daemon_name)
        # Embed the current tray icon into the launcher exe so the Windows
        # Startup apps entry (Task Manager / msconfig) shows the app icon
        # instead of the generic Python icon.
        if launcher != _sys.executable:
            _set_exe_icon(launcher, _get_tray_icon_image())
        ok, msg = install_task(launcher_exe=launcher, value_name=daemon_name)
        return jsonify({"ok": ok, "message": msg})

    @app.route("/api/daemon/uninstall-task", methods=["POST"])
    def api_daemon_uninstall_task():
        from windows_task import uninstall_task  # type: ignore
        daemon_name = Setting.get("daemon_name", "JobTrackerDaemon")
        ok, msg = uninstall_task(value_name=daemon_name)
        return jsonify({"ok": ok, "message": msg})

    # ── Tray icon helpers ────────────────────────────────────────────────────

    def _default_icon_image():
        """Generate the built-in blue-circle 'JT' icon (64×64 RGBA PIL Image)."""
        from PIL import Image, ImageDraw, ImageFont  # type: ignore
        size = 64
        img = Image.new("RGBA", (size, size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        draw.ellipse([2, 2, size - 2, size - 2], fill=(0, 102, 204, 255))
        try:
            font = ImageFont.truetype("arialbd.ttf", 26)
        except Exception:
            font = ImageFont.load_default()
        text = "JT"
        bbox = draw.textbbox((0, 0), text, font=font)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        draw.text(((size - tw) / 2, (size - th) / 2 - 2), text, fill="white", font=font)
        return img

    def _icon_to_png_bytes(img) -> bytes:
        """Convert a PIL Image to PNG bytes."""
        buf = BytesIO()
        img.convert("RGBA").save(buf, "PNG")
        return buf.getvalue()

    def _get_tray_icon_image():
        """Return the current tray icon as a 64×64 RGBA PIL Image."""
        from PIL import Image  # type: ignore
        icon_path = Setting.get("daemon_icon_path", "")
        icon_size = Setting.get("daemon_icon_size", "")
        img = None
        if icon_path and _os.path.isfile(icon_path):
            try:
                src = Image.open(icon_path)
                if hasattr(src, "ico") and src.format == "ICO":
                    sizes = src.ico.sizes()
                    target = (64, 64)
                    if icon_size:
                        try:
                            w, h = map(int, icon_size.split("x"))
                            target = (w, h)
                        except Exception:
                            pass
                    best = min(sizes, key=lambda s: abs(s[0] - target[0]) + abs(s[1] - target[1]))
                    src.size = best
                    src = src.copy()
                img = src.convert("RGBA").resize((64, 64), Image.LANCZOS)
            except Exception:
                img = None
        if img is None:
            img = _default_icon_image()
        return img

    def _set_exe_icon(exe_path: str, icon_image) -> None:
        """
        Embed icon_image as RT_ICON + RT_GROUP_ICON resources in exe_path so that
        the Windows Startup apps list shows the tray icon instead of the Python icon.
        Builds icons at 16, 32, 48, and 256 px using PNG-in-ICO format (supported
        by all modern Windows versions).  Fails gracefully if the exe is locked.
        """
        if _sys.platform != "win32":
            return
        try:
            import ctypes as _ct, struct as _st
            from PIL import Image as _PILImg  # type: ignore

            RT_ICON       = 3
            RT_GROUP_ICON = 14
            LANG_EN_US    = 0x0409

            # Render the icon at the four standard Windows icon sizes as PNG bytes
            SIZES = [16, 32, 48, 256]
            images: list = []
            for sz in SIZES:
                thumb = icon_image.convert("RGBA").resize((sz, sz), _PILImg.LANCZOS)
                _buf  = BytesIO()
                thumb.save(_buf, "PNG")
                images.append((sz, _buf.getvalue()))

            # Build the GRPICONDIR + GRPICONDIRENTRY array that goes into the
            # RT_GROUP_ICON resource.  Each entry mirrors ICONDIRENTRY but uses
            # a WORD nID (resource ID of the matching RT_ICON) instead of a
            # DWORD file offset.
            # Layout: BYTE bWidth, BYTE bHeight, BYTE bColorCount, BYTE bReserved,
            #         WORD wPlanes, WORD wBitCount, DWORD dwBytesInRes, WORD nID
            grp = _st.pack("<HHH", 0, 1, len(images))
            for i, (sz, data) in enumerate(images):
                grp += _st.pack(
                    "<BBBBHHIH",
                    sz if sz < 256 else 0,  # bWidth  (0 encodes 256)
                    sz if sz < 256 else 0,  # bHeight (0 encodes 256)
                    0,                       # bColorCount — 0 for 32-bit ARGB
                    0,                       # bReserved
                    1,                       # wPlanes
                    32,                      # wBitCount
                    len(data),               # dwBytesInRes
                    i + 1,                   # nID — matches the RT_ICON resource ID below
                )

            _k32 = _ct.WinDLL("kernel32", use_last_error=True)
            _k32.BeginUpdateResourceW.restype  = _ct.c_void_p
            _k32.BeginUpdateResourceW.argtypes = [_ct.c_wchar_p, _ct.c_bool]
            _k32.UpdateResourceW.restype  = _ct.c_bool
            _k32.UpdateResourceW.argtypes = [
                _ct.c_void_p, _ct.c_void_p, _ct.c_void_p,
                _ct.c_ushort, _ct.c_void_p, _ct.c_ulong,
            ]
            _k32.EndUpdateResourceW.restype  = _ct.c_bool
            _k32.EndUpdateResourceW.argtypes = [_ct.c_void_p, _ct.c_bool]

            hnd = _k32.BeginUpdateResourceW(exe_path, False)
            if not hnd:
                logger.warning(
                    "_set_exe_icon: BeginUpdateResourceW failed for '%s' (err=%d) — "
                    "exe may be in use; tray icon not embedded in Startup entry.",
                    exe_path, _ct.get_last_error(),
                )
                return

            ok = True
            # Write each individual icon image as RT_ICON
            for i, (_sz, data) in enumerate(images):
                arr = (_ct.c_char * len(data)).from_buffer_copy(data)
                if not _k32.UpdateResourceW(hnd, RT_ICON, i + 1, LANG_EN_US, arr, len(data)):
                    logger.warning(
                        "_set_exe_icon: UpdateResourceW RT_ICON[%d] failed (err=%d)",
                        i + 1, _ct.get_last_error(),
                    )
                    ok = False
                    break

            # Write the group icon directory as RT_GROUP_ICON (ID 1 = default app icon)
            if ok:
                grp_arr = (_ct.c_char * len(grp)).from_buffer_copy(grp)
                if not _k32.UpdateResourceW(hnd, RT_GROUP_ICON, 1, LANG_EN_US, grp_arr, len(grp)):
                    logger.warning(
                        "_set_exe_icon: UpdateResourceW RT_GROUP_ICON failed (err=%d)",
                        _ct.get_last_error(),
                    )
                    ok = False

            _k32.EndUpdateResourceW(hnd, not ok)  # False = commit, True = discard
            if ok:
                logger.info("_set_exe_icon: tray icon embedded in '%s'", exe_path)
        except ImportError:
            pass  # PIL not installed — icon embedding skipped silently
        except Exception:
            logger.exception("_set_exe_icon raised for '%s'", exe_path)

    @app.route("/api/daemon/icon-preview")
    def api_daemon_icon_preview():
        """Serve the current tray icon as a PNG (64×64)."""
        from PIL import Image  # type: ignore
        icon_path = Setting.get("daemon_icon_path", "")
        icon_size = Setting.get("daemon_icon_size", "")
        img = None
        if icon_path and _os.path.isfile(icon_path):
            try:
                src = Image.open(icon_path)
                # For ICO files pick the requested size or the largest available
                if hasattr(src, "ico") and src.format == "ICO":
                    sizes = src.ico.sizes()
                    target = (64, 64)
                    if icon_size:
                        try:
                            w, h = map(int, icon_size.split("x"))
                            target = (w, h)
                        except Exception:
                            pass
                    best = min(sizes, key=lambda s: abs(s[0] - target[0]) + abs(s[1] - target[1]))
                    src.size = best
                    src = src.copy()
                img = src.convert("RGBA").resize((64, 64), Image.LANCZOS)
            except Exception:
                img = None
        if img is None:
            img = _default_icon_image()
        return Response(_icon_to_png_bytes(img), mimetype="image/png",
                        headers={"Cache-Control": "no-store"})

    @app.route("/api/daemon/icon-upload", methods=["POST"])
    def api_daemon_icon_upload():
        """Accept a tray icon file upload; return list of available icon variants."""
        import base64
        from PIL import Image  # type: ignore
        f = request.files.get("icon")
        if not f or not f.filename:
            return jsonify({"ok": False, "error": "No file received"}), 400

        icons_dir = _os.path.join(_os.path.dirname(__file__), "icons")
        _os.makedirs(icons_dir, exist_ok=True)

        # Sanitise filename — keep only safe characters
        import re as _re
        safe_name = _re.sub(r"[^A-Za-z0-9_\-\.]", "_", _os.path.basename(f.filename or "icon"))
        save_path = _os.path.join(icons_dir, safe_name)
        f.save(save_path)

        # Extract all variants
        variants = []
        try:
            src = Image.open(save_path)
            if src.format == "ICO" and hasattr(src, "ico"):
                for sz in sorted(src.ico.sizes(), key=lambda s: s[0]):
                    try:
                        src.size = sz
                        frame = src.copy().convert("RGBA")
                        thumb = frame.resize((48, 48), Image.LANCZOS)
                        variants.append({
                            "size": f"{sz[0]}x{sz[1]}",
                            "data": "data:image/png;base64," + base64.b64encode(
                                _icon_to_png_bytes(thumb)).decode(),
                        })
                    except Exception:
                        pass
            if not variants:
                # Single image (PNG / JPG / BMP)
                img = src.convert("RGBA")
                w, h = img.size
                thumb = img.resize((48, 48), Image.LANCZOS)
                variants.append({
                    "size": f"{w}x{h}",
                    "data": "data:image/png;base64," + base64.b64encode(
                        _icon_to_png_bytes(thumb)).decode(),
                })
        except Exception as exc:
            _os.remove(save_path)
            return jsonify({"ok": False, "error": str(exc)}), 400

        return jsonify({"ok": True, "path": save_path, "variants": variants})

    @app.route("/api/daemon/icon-select", methods=["POST"])
    def api_daemon_icon_select():
        """Save the chosen icon path + size as settings."""
        data = request.get_json(silent=True) or {}
        path = data.get("path", "")
        size = data.get("size", "")
        Setting.set("daemon_icon_path", path)
        Setting.set("daemon_icon_size", size)
        return jsonify({"ok": True})

    @app.route("/api/daemon/icon-reset", methods=["POST"])
    def api_daemon_icon_reset():
        """Reset tray icon to the built-in default."""
        Setting.set("daemon_icon_path", "")
        Setting.set("daemon_icon_size", "")
        return jsonify({"ok": True})

    # ------------------------------------------------------------------
    # Response Email Accounts — IMAP inbox monitoring for application replies
    # ------------------------------------------------------------------

    @app.route("/api/response-emails/accounts")
    def api_response_email_accounts():
        """Return configured response-email accounts (passwords redacted)."""
        raw = Setting.get("response_email_accounts", "[]")
        try:
            accounts = json.loads(raw)
        except Exception:
            accounts = []
        safe = []
        for acc in accounts:
            a = dict(acc)
            a["password"] = "••••••••" if a.get("password") else ""
            safe.append(a)
        return jsonify(safe)

    @app.route("/api/response-emails/accounts/add", methods=["POST"])
    def api_response_email_add():
        """Add a new IMAP response-email account."""
        data = request.get_json(silent=True) or {}
        raw = Setting.get("response_email_accounts", "[]")
        try:
            accounts = json.loads(raw)
        except Exception:
            accounts = []
        new_account = {
            "id":           str(uuid.uuid4()),
            "label":        (data.get("label") or "").strip(),
            "email":        (data.get("email") or "").strip(),
            "imap_host":    (data.get("imap_host") or "").strip(),
            "imap_port":    int(data.get("imap_port") or 993),
            "use_ssl":      bool(data.get("use_ssl", True)),
            "password":     (data.get("password") or ""),
            "enabled":      True,
            "last_checked": None,
            "last_status":  None,
            "last_message": None,
        }
        accounts.append(new_account)
        Setting.set("response_email_accounts", json.dumps(accounts))
        safe = dict(new_account)
        safe["password"] = "••••••••" if safe.get("password") else ""
        return jsonify({"ok": True, "account": safe})

    @app.route("/api/response-emails/accounts/<account_id>/remove", methods=["POST"])
    def api_response_email_remove(account_id):
        """Delete a response-email account."""
        raw = Setting.get("response_email_accounts", "[]")
        try:
            accounts = json.loads(raw)
        except Exception:
            accounts = []
        accounts = [a for a in accounts if a.get("id") != account_id]
        Setting.set("response_email_accounts", json.dumps(accounts))
        return jsonify({"ok": True})

    @app.route("/api/response-emails/accounts/<account_id>/toggle", methods=["POST"])
    def api_response_email_toggle(account_id):
        """Enable or disable a response-email account."""
        raw = Setting.get("response_email_accounts", "[]")
        try:
            accounts = json.loads(raw)
        except Exception:
            accounts = []
        for acc in accounts:
            if acc.get("id") == account_id:
                acc["enabled"] = not acc.get("enabled", True)
                break
        Setting.set("response_email_accounts", json.dumps(accounts))
        return jsonify({"ok": True})

    @app.route("/api/response-emails/accounts/<account_id>/update-password", methods=["POST"])
    def api_response_email_update_password(account_id):
        """Update the stored IMAP password for an account."""
        data = request.get_json(silent=True) or {}
        password = data.get("password", "")
        raw = Setting.get("response_email_accounts", "[]")
        try:
            accounts = json.loads(raw)
        except Exception:
            accounts = []
        for acc in accounts:
            if acc.get("id") == account_id:
                acc["password"] = password
                break
        Setting.set("response_email_accounts", json.dumps(accounts))
        return jsonify({"ok": True})

    @app.route("/api/response-emails/test", methods=["POST"])
    def api_response_email_test():
        """Test an IMAP connection without saving. Returns ok + message."""
        data = request.get_json(silent=True) or {}
        from email_checker import test_connection
        ok, msg = test_connection(
            imap_host=(data.get("imap_host") or "").strip(),
            imap_port=int(data.get("imap_port") or 993),
            use_ssl=bool(data.get("use_ssl", True)),
            username=(data.get("email") or "").strip(),
            password=(data.get("password") or ""),
        )
        return jsonify({"ok": ok, "message": msg})

    @app.route("/api/response-emails/check-now", methods=["POST"])
    def api_response_email_check_now():
        """Manually trigger a full inbox check; streams progress via /check-status."""
        with _email_check_lock:
            if _email_check_state["running"]:
                return jsonify({"ok": False, "message": "Check already in progress."})
            _email_check_state["running"]  = True
            _email_check_state["finished"] = False
            _email_check_state["log"]      = []

        from email_checker import check_all_accounts
        _flask_app = app._get_current_object() if hasattr(app, "_get_current_object") else app

        def _progress(msg: str):
            with _email_check_lock:
                _email_check_state["log"].append(msg)

        def _run():
            try:
                check_all_accounts(_flask_app, progress_cb=_progress)
            except Exception as exc:
                _progress(f"Unexpected error: {exc}")
            finally:
                with _email_check_lock:
                    _email_check_state["running"]  = False
                    _email_check_state["finished"] = True

        threading.Thread(target=_run, daemon=True).start()
        return jsonify({"ok": True, "message": "Email check started."})

    @app.route("/api/response-emails/check-status")
    def api_response_email_check_status():
        """Return current progress state for the Check Now popup."""
        with _email_check_lock:
            return jsonify({
                "running":  _email_check_state["running"],
                "finished": _email_check_state["finished"],
                "log":      list(_email_check_state["log"]),
            })

    # ------------------------------------------------------------------
    # Email Review API
    # ------------------------------------------------------------------

    @app.route("/api/email-reviews/<int:review_id>/confirm", methods=["POST"])
    def api_confirm_email_review(review_id):
        """User confirms the suggested job match — apply status update + feedback."""
        from models import EmailReview, EmailMatchFeedback
        from email_checker import _sender_domain
        r = EmailReview.query.get_or_404(review_id)
        if r.review_status != "pending":
            return jsonify({"ok": False, "message": "Review already processed."})

        if r.suggested_app_id and r.suggested_app:
            a = r.suggested_app
            now_str = datetime.now().strftime("%Y-%m-%d")

            if r.classification == "declined":
                a.status = "rejected"
                notif_title = f"Declined: {a.job.title if a.job else 'Application'}"
                notif_msg   = f"{a.job.company if a.job else 'Company'} sent a decline. Status updated to Rejected."
            else:
                if a.status == "submitted":
                    a.status = "phone_screen"
                notif_title = f"Interview Invite: {a.job.title if a.job else 'Application'}"
                notif_msg   = f"{a.job.company if a.job else 'Company'} wants to schedule an interview!"

            auto_note = (
                f"\n\n[Confirmed {now_str}] {r.classification.title()} email "
                f"from {r.sender}. Saved to: {r.eml_path}"
            )
            a.notes = ((a.notes or "") + auto_note).strip()

            notif = Notification(
                type="follow_up",
                title=notif_title,
                message=notif_msg,
                job_id=a.job_id,
                application_id=a.id,
            )
            db.session.add(notif)

            # Positive feedback for learning
            s_domain = _sender_domain(r.sender)
            if s_domain:
                db.session.add(EmailMatchFeedback(
                    sender_domain=s_domain, app_id=a.id, is_confirmed=True
                ))

        r.review_status = "confirmed"
        db.session.commit()
        return jsonify({"ok": True, "message": "Match confirmed and application updated."})

    @app.route("/api/email-reviews/<int:review_id>/reject", methods=["POST"])
    def api_reject_email_review(review_id):
        """User says match is wrong — store negative feedback and try next best match."""
        from models import EmailReview, EmailMatchFeedback
        from email_checker import _match_application, _sender_domain
        r = EmailReview.query.get_or_404(review_id)
        if r.review_status != "pending":
            return jsonify({"ok": False, "message": "Review already processed."})

        s_domain = _sender_domain(r.sender)

        # Negative feedback for learning
        if s_domain and r.suggested_app_id:
            db.session.add(EmailMatchFeedback(
                sender_domain=s_domain, app_id=r.suggested_app_id, is_confirmed=False
            ))

        # Add current suggestion to rejected list
        rejected = r.get_rejected_ids()
        if r.suggested_app_id and r.suggested_app_id not in rejected:
            rejected.append(r.suggested_app_id)
        r.rejected_app_ids = json.dumps(rejected)
        db.session.commit()

        # Rebuild feedback sets (includes the new rejection just committed)
        confirmed_pairs = set()
        rejected_pairs  = set()
        for fb in EmailMatchFeedback.query.filter_by(sender_domain=s_domain).all():
            pair = (fb.sender_domain, fb.app_id)
            (confirmed_pairs if fb.is_confirmed else rejected_pairs).add(pair)

        active_apps = (
            Application.query
            .filter(Application.status.notin_(["draft", "withdrawn", "rejected"]))
            .all()
        )
        next_match = _match_application(
            s_domain, r.subject, r.body_preview or "", active_apps,
            exclude_app_ids=rejected,
            confirmed_pairs=confirmed_pairs,
            rejected_pairs=rejected_pairs,
        )

        r.suggested_app_id = next_match.id if next_match else None
        db.session.commit()

        if next_match and next_match.job:
            return jsonify({
                "ok": True, "has_new": True,
                "app": {
                    "id": next_match.id,
                    "job_id": next_match.job_id,
                    "status": next_match.status,
                    "status_label": next_match.status_label(),
                    "status_color": next_match.status_color(),
                    "job_title": next_match.job.title,
                    "job_company": next_match.job.company or "—",
                    "job_location": next_match.job.location or "—",
                    "job_url": next_match.job.url,
                    "job_salary": next_match.job.salary_range,
                    "job_description": (next_match.job.description or "")[:600],
                    "job_skills": next_match.job.get_matched_skills()[:8],
                    "job_score": next_match.job.match_score,
                }
            })
        return jsonify({"ok": True, "has_new": False,
                        "message": "No other matching jobs found for this email."})

    @app.route("/api/email-reviews/<int:review_id>/dismiss", methods=["POST"])
    def api_dismiss_email_review(review_id):
        """Dismiss the review without making any application changes."""
        from models import EmailReview
        r = EmailReview.query.get_or_404(review_id)
        r.review_status = "dismissed"
        db.session.commit()
        return jsonify({"ok": True})

    @app.route("/api/email-reviews/pending-count")
    def api_email_reviews_pending_count():
        from models import EmailReview
        count = EmailReview.query.filter_by(review_status="pending").count()
        return jsonify({"count": count})
